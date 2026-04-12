import base64
import hashlib
import html
import io
import json
import re
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import openpyxl
import pandas as pd
import streamlit as st


st.set_page_config(
    page_title="Sendeplan-Generator",
    page_icon="📦",
    layout="wide",
    initial_sidebar_state="collapsed",
)


# ============================================================
# DOMAIN-KONFIGURATION
# Zentrale Stelle für Depot-/Touren-Wissen, das sich ändern kann.
# ============================================================
WOCHENTAGE = {
    1: "Montag",
    2: "Dienstag",
    3: "Mittwoch",
    4: "Donnerstag",
    5: "Freitag",
    6: "Samstag",
}

# Keine Depot-Kategorien mehr – Sortiment-Zuordnung läuft nur noch über KSP-Schlüssel.

# Sortiment-Reihenfolge im Sendeplan
SORTIMENT_PRIO = {
    "fleisch- & wurst bedienung": 0,
    "fleisch- & wurst sb":        1,
    "heidemark":                  2,
}
SORTIMENT_ZUSATZ_KEYWORDS = ("avo", "werbemittel", "hamburger jungs", "lagerware", "divers")


def _sortiment_key(name: str) -> tuple:
    """Sortiment-Priorität: Fleisch/Heidemark zuerst, Zusatz-Kram zuletzt."""
    n = str(name).strip().lower()
    for key, prio in SORTIMENT_PRIO.items():
        if key in n:
            return (-1, prio)
    if any(k in n for k in SORTIMENT_ZUSATZ_KEYWORDS):
        return (1, 0)
    return (0, 0)

# Zusatz-Sortimente aus KSP Sheet.
# Spalte A(0) = Liefertag (1=Mo..6=Sa), B(1) = Tourname (Join-Key zu SAP.P).
# Danach je 3 Spalten pro Sortiment: Name | Uhrzeit | Bestelltag
# C/D/E = Lagerware, F/G/H = AVO, I/J/K = WM Sonder, L/M/N = WM, O/P/Q = HJ, R/S/T = Divers
# Nur hinzufügen wenn Uhrzeit UND Tag vorhanden.
KST_ZUSATZ_GRUPPEN = [
    (2,  "Lagerware"),
    (5,  "AVO"),
    (8,  "Werbemittel Sonder"),
    (11, "Werbemittel"),
    (14, "Hamburger Jungs"),
    (17, "Divers"),
]

TAG_ABKUERZUNGEN = {
    "mo": "Montag", "die": "Dienstag", "mitt": "Mittwoch", "mi": "Mittwoch",
    "don": "Donnerstag", "do": "Donnerstag", "fr": "Freitag", "sa": "Samstag", "so": "Sonntag",
}

UPLOAD_CONFIG = {
    "kunden": {
        "label": "Kundenliste hochladen",
        "help": "Verwendet feste Excel-Spalten: A, I, J, K, L, M, N",
        "mapping": {
            "Fachberater": "A",
            "CSB_Nr": "I",
            "SAP_Nr": "J",
            "Name": "K",
            "Strasse": "L",
            "PLZ": "M",
            "Ort": "N",
        },
        "required": ["Fachberater", "CSB_Nr", "SAP_Nr", "Name", "Strasse", "PLZ", "Ort"],
        "key": "SAP_Nr",
    },
    "sap": {
        "label": "SAP-Datei hochladen",
        "help": "Verwendet feste Excel-Spalten: A, G, H, I, O, P, Y",
        "mapping": {
            "SAP_Nr": "A",
            "Warengruppe": "C",
            "Liefertag_Raw": "G",
            "Bestelltag": "H",
            "Bestellzeitende": "I",
            "Liefertyp_ID": "O",
            "KSP_Schluessel": "P",
            "Rahmentour_Raw": "Y",
        },
        "required": ["SAP_Nr", "Liefertag_Raw", "Bestelltag", "Bestellzeitende", "Liefertyp_ID", "KSP_Schluessel", "Rahmentour_Raw"],
        "key": "SAP_Nr",
    },
    "transport": {
        "label": "Transportgruppen hochladen",
        "help": "Verwendet feste Excel-Spalten: A, C",
        "mapping": {
            "Liefertyp_ID": "A",
            "Liefertyp_Name": "C",
        },
        "required": ["Liefertyp_ID", "Liefertyp_Name"],
        "key": "Liefertyp_ID",
    },
    "anfahrt": {
        "label": "Anfahrtshinweise hochladen (optional)",
        "help": "CSV: A=Rahmentour, B=CSB-Nr (Matchkey), G=Behälter, H=Anfahrtshinweis, K=Foto",
        "mapping": {
            "SAP_Nr_CSV":     "A",
            "CSB_Nr_CSV":     "B",
            "Behaelter":      "G",
            "Anfahrtshinweis": "H",
            "Foto":           "K",
        },
        "required": ["CSB_Nr_CSV"],
        "key": "CSB_Nr_CSV",
    },
}

# (Kostenstellenplan wird nur noch für Zusatz-Sortimente genutzt – erstes Sheet.)


# ============================================================
# HILFSFUNKTIONEN
# ============================================================
def normalize_text(value) -> str:
    if pd.isna(value):
        return ""
    return str(value).strip()


def normalize_digits(value) -> str:
    text = normalize_text(value)
    digits = "".join(ch for ch in text if ch.isdigit())
    return digits if digits else text


def _normalize_ksp_key(value) -> str:
    """Normalisiert KSP-Schlüssel: 1.0 -> '1', '1.0' -> '1', 1 -> '1'."""
    text = normalize_text(value)
    # Float-artige Strings: '1.0' -> '1'
    if re.match(r'^\d+\.0$', text):
        text = text[:-2]
    return text


def day_name_from_number(value) -> str:
    try:
        return WOCHENTAGE.get(int(str(value).strip()), "Unbekannt")
    except (TypeError, ValueError):
        return "Unbekannt"


def validate_required_columns(df: pd.DataFrame, required_columns: List[str], name: str) -> None:
    missing = [col for col in required_columns if col not in df.columns]
    if missing:
        raise ValueError(f"{name} fehlt Pflichtspalten: {', '.join(missing)}")


def excel_column_to_index(column_letter: str) -> int:
    result = 0
    for char in column_letter.upper():
        if not char.isalpha():
            raise ValueError(f"Ungültiger Spaltenbuchstabe: {column_letter}")
        result = result * 26 + (ord(char) - ord("A") + 1)
    return result - 1


def read_upload_to_raw_dataframe(file_bytes: bytes, filename: str, csv_separator: str) -> pd.DataFrame:
    suffix = Path(filename).suffix.lower()

    if suffix == ".csv":
        errors = []
        for encoding in ["utf-8-sig", "utf-8", "latin1", "cp1252"]:
            try:
                return pd.read_csv(
                    io.BytesIO(file_bytes),
                    sep=csv_separator,
                    header=None,
                    dtype=str,
                    encoding=encoding,
                    keep_default_na=False,
                )
            except Exception as exc:
                errors.append(f"{encoding}: {exc}")
        raise ValueError(f"CSV konnte nicht gelesen werden. Versuchte Encodings: {' | '.join(errors)}")

    if suffix in {".xlsx", ".xls", ".xlsm"}:
        return pd.read_excel(
            io.BytesIO(file_bytes),
            header=None,
            dtype=str,
            keep_default_na=False,
        )

    raise ValueError(f"Nicht unterstütztes Dateiformat: {suffix}")


def extract_columns_by_letter(raw_df: pd.DataFrame, mapping: Dict[str, str], dataset_name: str) -> pd.DataFrame:
    extracted: Dict[str, pd.Series] = {}
    for target_name, column_letter in mapping.items():
        column_index = excel_column_to_index(column_letter)
        if column_index >= raw_df.shape[1]:
            raise ValueError(
                f"{dataset_name}: benötigte Spalte {column_letter} ist nicht vorhanden. "
                f"Gefundene Spaltenanzahl: {raw_df.shape[1]}"
            )
        extracted[target_name] = raw_df.iloc[:, column_index]
    return pd.DataFrame(extracted)


def cleanup_dataframe(df: pd.DataFrame, key_column: str) -> pd.DataFrame:
    result = df.copy()
    for column in result.columns:
        result[column] = result[column].map(normalize_text)

    result = result.replace("", pd.NA).dropna(how="all").fillna("")
    result = result[result[key_column].map(normalize_text) != ""].copy()

    if not result.empty:
        first_key = normalize_text(result.iloc[0][key_column]).lower()
        header_like_tokens = {
            key_column.lower(),
            key_column.lower().replace("_", " "),
            "sap",
            "sap-nr",
            "sap nr",
            "liefertyp_id",
            "liefertyp id",
            "sap rahmentour",
            "sap_von",
            "sap von",
            "csb tournummer",
            "kundennummer",
            "kundennummer / markt",
            "kunden-nr",
            "kunden nr",
        }
        # Auch Teilstring-Match: wenn der erste Schlüssel mit einem Token beginnt
        is_header = first_key in header_like_tokens or any(
            first_key.startswith(t) for t in header_like_tokens if len(t) > 3
        )
        if is_header:
            result = result.iloc[1:].copy()

    return result.reset_index(drop=True)


def load_structured_upload(file_bytes: bytes, filename: str, csv_separator: str, dataset_key: str) -> pd.DataFrame:
    config = UPLOAD_CONFIG[dataset_key]
    raw_df = read_upload_to_raw_dataframe(file_bytes, filename, csv_separator)
    structured_df = extract_columns_by_letter(raw_df, config["mapping"], config["label"])
    structured_df = cleanup_dataframe(structured_df, config["key"])
    validate_required_columns(structured_df, config["required"], config["label"])
    return structured_df


# ============================================================
# ZUSATZ-SORTIMENTE AUS KOSTENSTELLENPLAN (AVO, WERBEMITTEL …)
# ============================================================

# Zuordnung über KSP-Schlüssel (SAP Spalte P → KSP Spalte B) + Liefertag

def _parse_kst_time(val) -> str:
    """Wandelt verschiedene Zeit-Formate in 'HH:MM'.

    Erkennt: 915 -> '09:15', 2000 -> '20:00', '9:00' -> '09:00',
    '20:00' -> '20:00', datetime.time(9,0) -> '09:00'.
    """
    if val is None:
        return ""
    # datetime.time Objekt (kommt aus openpyxl bei formatierten Zellen)
    if hasattr(val, "hour") and hasattr(val, "minute"):
        return f"{val.hour:02d}:{val.minute:02d}"
    text = str(val).strip()
    if not text:
        return ""
    # Bereits im Format "H:MM" oder "HH:MM"?
    m = re.match(r'^(\d{1,2}):(\d{2})$', text)
    if m:
        return f"{int(m.group(1)):02d}:{m.group(2)}"
    # Numerisch: 900 -> '09:00', 2000 -> '20:00'
    try:
        n = int(float(text))
    except (ValueError, TypeError):
        return ""
    if n <= 0:
        return ""
    s = f"{n:04d}"
    return f"{s[:2]}:{s[2:]}"


def _parse_kst_tag(val) -> str:
    """Wandelt 'Don' -> 'Donnerstag', 'Fr' -> 'Freitag' etc."""
    if val is None:
        return ""
    key = str(val).strip().lower()
    return TAG_ABKUERZUNGEN.get(key, str(val).strip())


def extract_zusatz_schedule(file_bytes: bytes, filename: str) -> pd.DataFrame:
    """Extrahiert den Bestellplan fuer Zusatz-Sortimente (AVO, Werbemittel, …)
    aus dem Kostenstellenplan (erstes Sheet).

    Flaches Layout – jede Zeile ist eine Tour:
      A = Liefertag (1=Mo … 6=Sa)
      B = KSP-Schlüssel (Join-Key zu SAP Spalte P)
      Dann je 3 Spalten pro Sortiment: Name | Uhrzeit | Bestelltag
        C/D/E = Lagerware, F/G/H = AVO, I/J/K = WM Sonder,
        L/M/N = Werbemittel, O/P/Q = Hamburger Jungs

    Sortiment wird NUR erzeugt wenn BEIDE (Uhrzeit + Tag) gefüllt sind.

    Ergebnis-DataFrame: ksp_schluessel | liefertag | sortiment | bestelltag | bestellzeitende
    """
    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True)
    except Exception as exc:
        raise ValueError(
            f"Kostenstellenplan konnte nicht als Excel gelesen werden ({filename}): {exc}"
        ) from exc
    # Erstes Sheet verwenden (heißt je nach Datei "Tabelle1", "CSB Standard" o.ä.)
    ws = wb[wb.sheetnames[0]]
    all_rows = list(ws.iter_rows(values_only=True))
    records = []

    for row in all_rows:
        if not row or len(row) < 3:
            continue

        # Spalte A = Liefertag als Zahl (1=Mo … 6=Sa)
        a_raw = row[0]
        try:
            liefertag_num = int(float(str(a_raw).strip()))
        except (ValueError, TypeError):
            continue  # Kopf-/Leerzeile
        liefertag_name = WOCHENTAGE.get(liefertag_num)
        if not liefertag_name:
            continue  # ungültige Zahl

        # Spalte B = KSP-Schlüssel
        b = row[1] if len(row) > 1 else None
        if b is None:
            continue
        ksp_key = _normalize_ksp_key(b)
        if not ksp_key:
            continue

        # Für jede Zusatz-Gruppe: (col_start, sortiment_name)
        for col_start, sortiment_name in KST_ZUSATZ_GRUPPEN:
            time_col = col_start + 1
            day_col  = col_start + 2
            if len(row) <= day_col:
                continue

            zeit_val = row[time_col]
            tag_val  = row[day_col]

            # NUR wenn BEIDE gefüllt
            bestellzeitende = _parse_kst_time(zeit_val)
            bestelltag      = _parse_kst_tag(tag_val)

            if not bestellzeitende or not bestelltag:
                continue

            records.append({
                "ksp_schluessel":  ksp_key,
                "liefertag":       liefertag_name,
                "sortiment":       sortiment_name,
                "bestelltag":      bestelltag,
                "bestellzeitende": bestellzeitende,
            })

    return pd.DataFrame(records) if records else pd.DataFrame(
        columns=["ksp_schluessel","liefertag","sortiment","bestelltag","bestellzeitende"]
    )


def build_zusatz_plan_rows(plan_rows: pd.DataFrame, zusatz_schedule: pd.DataFrame) -> pd.DataFrame:
    """Generiert synthetische Planzeilen fuer AVO, Werbemittel etc.

    Fuer jede einzigartige (SAP_Nr, Liefertag) Kombination in plan_rows wird geprueft,
    ob es passende Eintraege in zusatz_schedule gibt (via KSP_Schluessel x Liefertag).
    Falls ja, wird eine neue Zeile erzeugt und angehaengt.
    """
    if zusatz_schedule.empty or plan_rows.empty:
        return plan_rows

    if "KSP_Schluessel" not in plan_rows.columns:
        return plan_rows

    # Basis-Info pro (SAP_Nr, Liefertag): nimm erste Zeile
    basis_cols = ["SAP_Nr", "Liefertag", "KSP_Schluessel",
                  "Rahmentour_Raw",
                  "Bestelltag", "SortKey_Bestelltag",
                  "CSB_Nr", "Name", "Strasse", "PLZ", "Ort", "Fachberater",
                  "Liefertyp_ID", "Liefertyp_Name"]
    avail_cols = [c for c in basis_cols if c in plan_rows.columns]

    basis = (
        plan_rows[avail_cols]
        .drop_duplicates(subset=["SAP_Nr", "Liefertag"])
        .copy()
    )

    # Normalize für Merge – _normalize_ksp_key entfernt z.B. ".0" bei float-Werten
    sched = zusatz_schedule.copy()
    sched["_ksp_norm"] = sched["ksp_schluessel"].map(_normalize_ksp_key).str.lower()
    sched["_lt_norm"]  = sched["liefertag"].str.strip().str.lower()
    basis["_ksp_norm"] = basis["KSP_Schluessel"].map(_normalize_ksp_key).str.lower()
    basis["_lt_norm"]  = basis["Liefertag"].str.strip().str.lower()

    # Leere KSP-Schlüssel / Liefertage ausfiltern
    basis = basis[(basis["_ksp_norm"] != "") & (basis["_lt_norm"] != "")]

    if basis.empty:
        return plan_rows

    # Merge über KSP_Schluessel + Liefertag
    merged = basis.merge(sched, on=["_ksp_norm", "_lt_norm"], how="inner")

    if merged.empty:
        return plan_rows

    # Zusatz-Spalten setzen
    merged["Sortiment"] = merged["sortiment"]
    merged["Bestelltag_Name"] = merged["bestelltag"]
    merged["Bestellzeitende"] = merged["bestellzeitende"]
    merged["SortKey_Sortiment"] = merged["sortiment"].map(lambda n: (1, 0))
    merged["_ist_zusatz"] = True
    merged["Liefertyp_ID"] = ""

    # Aufräumen: nur plan_rows-Spalten behalten, Rest auffüllen
    drop_cols = ["_ksp_norm", "_lt_norm", "ksp_schluessel", "liefertag", "sortiment",
                 "bestelltag", "bestellzeitende"]
    merged = merged.drop(columns=[c for c in drop_cols if c in merged.columns], errors="ignore")

    for col in plan_rows.columns:
        if col not in merged.columns:
            merged[col] = ""

    combined = pd.concat([plan_rows, merged[plan_rows.columns]], ignore_index=True)
    # _ist_zusatz markiert KSP-Zeilen – wird bei merged[plan_rows.columns] abgeschnitten,
    # daher hier nachtragen: neue Zeilen (ab len(plan_rows)) sind KSP.
    combined["_ist_zusatz"] = False
    combined.loc[len(plan_rows):, "_ist_zusatz"] = True
    return combined


def prepare_dataframes(
    kunden_bytes: bytes,
    kunden_name: str,
    sap_bytes: bytes,
    sap_name: str,
    transport_bytes: bytes,
    transport_name: str,
    kostenstellen_bytes: bytes,
    kostenstellen_name: str,
    csv_separator: str,
    anfahrt_bytes: Optional[bytes] = None,
    anfahrt_name: Optional[str] = None,
) -> Tuple[pd.DataFrame, pd.DataFrame, Dict[str, int], pd.DataFrame]:
    df_kunden = load_structured_upload(kunden_bytes, kunden_name, csv_separator, "kunden")
    df_sap = load_structured_upload(sap_bytes, sap_name, csv_separator, "sap")
    # Nur Zeilen mit Warengruppe FLEISCH (Spalte C) übernehmen
    df_sap = df_sap[df_sap["Warengruppe"].str.upper().str.strip() == "FLEISCH"].copy()
    df_sap = df_sap.drop(columns=["Warengruppe"])

    # Jede FLEISCH-Zeile MUSS einen KSP-Schlüssel haben – Zeilen ohne werden entfernt + Warnung
    _ksp_missing = df_sap["KSP_Schluessel"].isna() | (df_sap["KSP_Schluessel"].astype(str).str.strip() == "")
    if _ksp_missing.any():
        _n_miss = int(_ksp_missing.sum())
        _sample = df_sap.loc[_ksp_missing, "SAP_Nr"].head(10).tolist()
        st.warning(
            f"{_n_miss} FLEISCH-Zeile(n) ohne KSP-Schlüssel entfernt. "
            f"Betroffene SAP-Nr (max. 10): {', '.join(str(s) for s in _sample)}"
        )
        df_sap = df_sap[~_ksp_missing].copy()
    df_transport = load_structured_upload(transport_bytes, transport_name, csv_separator, "transport")

    # normalize_text wurde bereits in cleanup_dataframe() angewendet – kein zweiter Pass nötig.

    df_sap["Bestelltag_Name"] = df_sap["Bestelltag"].map(day_name_from_number)

    df_sap = df_sap.merge(df_transport, on="Liefertyp_ID", how="left")

    # Echte Duplikate aus SAP entfernen: gleiche SAP + Liefertag + Bestelltag + Sortiment.
    df_sap = df_sap.drop_duplicates(
        subset=["SAP_Nr", "Liefertag_Raw", "Bestelltag", "Liefertyp_ID"], keep="first"
    ).copy()

    kunden_basis = df_kunden.merge(
        df_sap[["SAP_Nr", "Rahmentour_Raw"]].drop_duplicates(subset=["SAP_Nr"]),
        on="SAP_Nr",
        how="left",
    )

    # Anfahrtshinweise aus optionalem CSV einbinden (Join via CSB_Nr)
    _anfahrt_cols = ["SAP_Nr_CSV", "Behaelter", "Anfahrtshinweis", "Foto"]
    if anfahrt_bytes and anfahrt_name:
        try:
            df_anfahrt = load_structured_upload(anfahrt_bytes, anfahrt_name, csv_separator, "anfahrt")
            df_anfahrt = df_anfahrt.rename(columns={"CSB_Nr_CSV": "CSB_Nr"})
            df_anfahrt["CSB_Nr"] = df_anfahrt["CSB_Nr"].map(normalize_digits)
            kunden_basis["CSB_Nr"] = kunden_basis["CSB_Nr"].map(normalize_digits)
            kunden_basis = kunden_basis.merge(
                df_anfahrt[["CSB_Nr"] + _anfahrt_cols].drop_duplicates("CSB_Nr"),
                on="CSB_Nr",
                how="left",
            )
        except Exception as exc:
            st.warning(f"Anfahrtshinweise konnten nicht geladen werden: {exc}")
            for col in _anfahrt_cols:
                kunden_basis[col] = ""
    else:
        for col in _anfahrt_cols:
            kunden_basis[col] = ""

    kunden_basis[_anfahrt_cols] = kunden_basis[_anfahrt_cols].fillna("")

    # Basis-Merge: Kundenstamm-Spalten an plan_rows anhängen
    plan_rows = df_sap.merge(
        kunden_basis[["SAP_Nr", "CSB_Nr", "Name", "Strasse", "PLZ", "Ort", "Fachberater"]],
        on="SAP_Nr",
        how="left",
    )

    # Liefertag vectorisiert aus SAP Spalte G (1=Mo … 6=Sa)
    plan_rows["Liefertag"] = (
        plan_rows["Liefertag_Raw"]
        .map(normalize_digits)
        .str[:1]
        .map(lambda d: WOCHENTAGE.get(int(d), "Unbekannt") if d.isdigit() else "Unbekannt")
    )
    plan_rows["Sortiment"] = plan_rows["Liefertyp_Name"].fillna("")
    plan_rows["Bestellzeitende"] = plan_rows["Bestellzeitende"].fillna("")
    plan_rows["SortKey_Bestelltag"] = pd.to_numeric(plan_rows["Bestelltag"], errors="coerce").fillna(99)
    plan_rows["SortKey_Sortiment"] = plan_rows["Sortiment"].fillna("").map(_sortiment_key)

    # Zusatz-Sortimente (AVO, Werbemittel etc.) aus Kostenstellenplan generieren
    # Join über KSP_Schluessel (SAP Spalte P) + Liefertag
    zusatz_schedule = extract_zusatz_schedule(kostenstellen_bytes, kostenstellen_name)
    plan_rows = build_zusatz_plan_rows(plan_rows, zusatz_schedule)

    counts = {"Alle": int(len(kunden_basis))}

    # Vorberechnete Suchspalte für schnelle filter_customers-Aufrufe
    kunden_basis["_search_blob"] = (
        kunden_basis["SAP_Nr"].fillna("") + " " +
        kunden_basis["Name"].fillna("") + " " +
        kunden_basis["CSB_Nr"].fillna("") + " " +
        kunden_basis["Ort"].fillna("")
    ).str.lower()

    return kunden_basis, plan_rows, counts, df_sap


# ============================================================
# DEBUG / QUALITÄTSPRÜFUNG
# ============================================================
def build_debug_report(
    plan_rows: pd.DataFrame,
    df_sap_raw: pd.DataFrame,
) -> Dict[str, pd.DataFrame]:
    """Erstellt Qualitäts-Reports für SAP-Daten."""
    reports: Dict[str, pd.DataFrame] = {}

    def safe_cols(df: pd.DataFrame, cols: List[str]) -> pd.DataFrame:
        """Nur Spalten auswählen die wirklich vorhanden sind."""
        return df[[c for c in cols if c in df.columns]]

    # Kunden ohne Zusatz-Sortimente
    if "_ist_zusatz" in plan_rows.columns:
        sap_with_zusatz = set(plan_rows.loc[plan_rows["_ist_zusatz"] == True, "SAP_Nr"].unique())
        sap_all = set(plan_rows["SAP_Nr"].unique())
        sap_without = sap_all - sap_with_zusatz
        ohne_zusatz = plan_rows[plan_rows["SAP_Nr"].isin(sap_without)]
        reports["Ohne Zusatz-Sortimente"] = safe_cols(
            ohne_zusatz,
            ["SAP_Nr", "Name", "KSP_Schluessel", "Liefertag", "Sortiment"]
        ).drop_duplicates(subset=["SAP_Nr"]).reset_index(drop=True)
    else:
        reports["Ohne Zusatz-Sortimente"] = pd.DataFrame()

    return reports


# ============================================================
# MASSENDRUCK – STANDARDWOCHE & SORTIERLOGIK
# ============================================================

def render_debug_tab(reports: Dict[str, pd.DataFrame]) -> None:
    """Zeigt Debug-Reports im Streamlit-Tab."""
    total_issues = sum(len(df) for df in reports.values())
    if total_issues == 0:
        st.success("✅ Keine Auffälligkeiten gefunden.")
        return

    # Gesamt-Export aller Reports als Excel (ein Sheet pro Report)
    _buf = io.BytesIO()
    with pd.ExcelWriter(_buf, engine="openpyxl") as _writer:
        for _title, _df in reports.items():
            if not _df.empty:
                _sheet = _title[:31]  # Excel-Sheetname max 31 Zeichen
                _df.to_excel(_writer, sheet_name=_sheet, index=False)
    _buf.seek(0)
    st.download_button(
        label="📥 Alle Debug-Daten als Excel herunterladen",
        data=_buf.getvalue(),
        file_name="sendeplan_debug.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )
    st.divider()

    for title, df in reports.items():
        count = len(df)
        icon = "✅" if count == 0 else "⚠️"
        with st.expander(f"{icon} {title} ({count} Einträge)", expanded=count > 0):
            if df.empty:
                st.success("Keine Einträge.")
            else:
                st.dataframe(df, use_container_width=True, hide_index=True)
                # Download für diesen Report
                _buf2 = io.BytesIO()
                df.to_excel(_buf2, index=False, engine="openpyxl")
                _buf2.seek(0)
                _safe = title.replace("/", "-").replace(" ", "_")
                st.download_button(
                    label=f"📥 {title} exportieren",
                    data=_buf2.getvalue(),
                    file_name=f"debug_{_safe}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key=f"dl_{_safe}",
                )


# ============================================================
# FILTER
# ============================================================
def filter_customers(df_customers: pd.DataFrame, search_text: str) -> pd.DataFrame:
    mask = pd.Series(True, index=df_customers.index)

    search = normalize_text(search_text).lower()
    if search:
        if "_search_blob" in df_customers.columns:
            mask &= df_customers["_search_blob"].str.contains(search, na=False)
        else:
            mask &= (
                df_customers["SAP_Nr"].str.lower().str.contains(search, na=False)
                | df_customers["Name"].str.lower().str.contains(search, na=False)
                | df_customers["CSB_Nr"].str.lower().str.contains(search, na=False)
            )

    return df_customers.loc[mask].sort_values(["Name", "SAP_Nr"], na_position="last").reset_index(drop=True)


# ============================================================
# STREAMLIT-LAYOUT
# ============================================================
def streamlit_css() -> str:
    return """
    <style>
        .status-ok   { color: #3fb950; font-size: 0.85rem; }
        .status-miss { color: #f85149; font-size: 0.85rem; }
    </style>
    """


# ============================================================
# HTML EXPORT UND CSS
# ============================================================
def export_css() -> str:
    return """
    <style>
        :root {
            --ink:        #111;
            --ink-soft:   #444;
            --ink-muted:  #7a8fa0;
            --bg-main:    #111b25;
            --bg-card:    #18273a;
            --bg-hover:   #1e3347;
            --accent:     #f0a500;
            --accent-dim: #7a5200;
            --accent-soft:#fff3d0;
            --red:        #c00;
            --green:      #1a9e52;
            --border:     rgba(255,255,255,0.07);
            --paper-bg:   #ffffff;
        }

        *, *::before, *::after { box-sizing: border-box; margin: 0; padding: 0; }

        body {
            background: var(--bg-main);
            font-family: 'Segoe UI', system-ui, -apple-system, sans-serif;
            font-size: 10pt;
            color: var(--ink);
            display: flex;
            flex-direction: row;
            min-height: 100vh;
            margin: 0;
        }

        /* ══════════════════════════════════════
           SIDEBAR – Helles Design
        ══════════════════════════════════════ */
        .sidebar {
            --sb-bg:        #f5f7fa;
            --sb-border:    #dde2ea;
            --sb-hover:     #e8ecf2;
            --sb-active:    #e6a100;
            --sb-text:      #1a2332;
            --sb-muted:     #6b7a90;
            --sb-input-bg:  #ffffff;
            --sb-shadow:    rgba(0,0,0,0.06);

            width: 240px;
            min-width: 240px;
            background: var(--sb-bg);
            border-right: 1px solid var(--sb-border);
            box-shadow: 2px 0 12px var(--sb-shadow);
            min-height: 100vh;
            position: sticky;
            top: 0;
            height: 100vh;
            display: flex;
            flex-direction: column;
            z-index: 100;
            overflow-y: auto;
            scrollbar-width: thin;
            scrollbar-color: var(--sb-border) transparent;
        }
        .sidebar-logo {
            padding: 16px 14px 14px;
            border-bottom: 1px solid var(--sb-border);
            display: flex;
            align-items: center;
            gap: 12px;
            background: #fff;
        }
        .sidebar-logo-icon {
            width: 48px; height: 48px;
            background: var(--sb-active);
            border-radius: 10px;
            display: flex; align-items: center; justify-content: center;
            font-size: 22px; flex-shrink: 0;
            box-shadow: 0 2px 8px rgba(230,161,0,0.35);
        }
        .sidebar-logo-text {
            font-size: 12px;
            font-weight: 800;
            color: #1a2332;
            letter-spacing: 0.01em;
            line-height: 1.25;
        }
        .sidebar-logo-sub {
            font-size: 10.5px;
            color: var(--sb-active);
            font-weight: 600;
            margin-top: 2px;
            letter-spacing: 0.02em;
        }
        .sidebar-section {
            padding: 14px 12px 10px;
            border-bottom: 1px solid var(--sb-border);
        }
        .sidebar-label {
            font-size: 9px;
            font-weight: 700;
            color: var(--sb-muted);
            text-transform: uppercase;
            letter-spacing: 0.13em;
            margin-bottom: 8px;
        }
        .sidebar input[type=text] {
            width: 100%;
            border: 1.5px solid var(--sb-border);
            border-radius: 8px;
            padding: 8px 11px;
            font-size: 12px;
            font-family: inherit;
            outline: none;
            background: var(--sb-input-bg);
            color: var(--sb-text);
            transition: border-color 0.15s, box-shadow 0.15s;
            box-shadow: 0 1px 3px rgba(0,0,0,0.05) inset;
        }
        .sidebar input[type=text]::placeholder { color: #aab2be; }
        .sidebar input[type=text]:focus {
            border-color: var(--sb-active);
            box-shadow: 0 0 0 3px rgba(230,161,0,0.15);
        }
        .search-btn {
            border: 1.5px solid var(--sb-border);
            border-radius: 7px;
            padding: 7px 11px;
            font-size: 12px;
            font-weight: 600;
            font-family: inherit;
            cursor: pointer;
            background: #fff;
            color: var(--sb-text);
            transition: all 0.14s;
        }
        .search-btn:hover { background: var(--sb-hover); border-color: #c5cdd8; }
        .search-btn.reset {
            background: #fff0f0;
            border-color: #f5b8b8;
            color: #c0392b;
        }
        .search-btn.reset:hover { background: #fde8e8; border-color: #e8a0a0; }
        .search-nav-row {
            display: flex;
            gap: 5px;
            margin-top: 8px;
            align-items: center;
        }
        .search-count {
            font-size: 11px;
            font-family: 'Courier New', monospace;
            color: var(--sb-muted);
            flex: 1;
        }
        .search-empty {
            display: none;
            background: #fff8e6;
            color: #9a6800;
            border-radius: 6px;
            padding: 6px 10px;
            font-size: 11px;
            font-weight: 600;
            margin-top: 6px;
            border: 1px solid #f0c84a;
        }
        .sidebar-print-btn {
            display: block;
            width: calc(100% - 28px);
            margin: 14px;
            border: none;
            border-radius: 10px;
            padding: 11px;
            font-size: 13px;
            font-weight: 700;
            font-family: inherit;
            cursor: pointer;
            background: var(--sb-active);
            color: #fff;
            text-align: center;
            transition: all 0.15s;
            letter-spacing: 0.01em;
            box-shadow: 0 2px 10px rgba(230,161,0,0.3);
        }
        .sidebar-print-btn:hover {
            background: #f5b400;
            transform: translateY(-1px);
            box-shadow: 0 4px 16px rgba(230,161,0,0.4);
        }
        .sidebar-subtitle-group {
            padding: 14px 12px 8px;
        }

        /* ══════════════════════════════════════
           MAIN + PAGE-STACK
        ══════════════════════════════════════ */
        .main-content { flex: 1; min-width: 0; }
        .page-stack { padding: 28px 0 40px; }

        /* ══════════════════════════════════════
           A4-PAPIER
        ══════════════════════════════════════ */
        .paper {
            width: 210mm;
            height: 297mm;
            overflow: hidden;
            margin: 0 auto 28px auto;
            background: var(--paper-bg);
            box-shadow:
                0 2px 4px rgba(0,0,0,0.12),
                0 8px 32px rgba(0,0,0,0.28);
            padding: 0;
            position: relative;
            border-radius: 3px;
        }
        .paper::before { display: none; }
        .paper-inner {
            width: 210mm;
            padding: 14mm 15mm 12mm 15mm;
            box-sizing: border-box;
            transform-origin: top left;
            zoom: 1;
        }

        /* ══════════════════════════════════════
           SEITENHEADER
        ══════════════════════════════════════ */
        .doc-header {
            display: grid;
            grid-template-columns: 52mm 1fr 44mm;
            gap: 3mm;
            align-items: flex-start;
            margin-bottom: 2mm;
            padding-bottom: 0;
            border-bottom: none;
        }
        .doc-address {
            font-size: 9pt;
            line-height: 1.5;
            color: #333;
        }
        .doc-address strong {
            font-size: 9.5pt;
            font-weight: 700;
            display: block;
            margin-bottom: 0.5mm;
            color: #111;
        }
        .doc-title-block { text-align: center; padding: 0 2mm; }
        .doc-title {
            font-size: 16pt;
            font-weight: 700;
            line-height: 1.15;
            margin-bottom: 1.5mm;
            color: #111;
        }
        .doc-subtitle {
            font-size: 22pt;
            font-weight: 700;
            color: #c00;
            margin-bottom: 1mm;
            cursor: text;
            border-radius: 3px;
            padding: 1px 4px;
            outline: none;
            display: inline-block;
        }
        .doc-subtitle:hover { background: rgba(200,0,0,0.06); }
        .doc-subtitle:focus {
            background: rgba(200,0,0,0.08);
            box-shadow: 0 0 0 2px rgba(200,0,0,0.15);
        }
        @media print {
            .doc-subtitle:hover, .doc-subtitle:focus {
                background: none; box-shadow: none;
            }
        }
        .doc-allsortiments { font-size: 8pt; color: #666; }
        .doc-logo { text-align: right; }

        /* ══════════════════════════════════════
           INFOLEISTE
        ══════════════════════════════════════ */
        .doc-infobar {
            font-size: 9pt;
            margin: 2mm 0 2mm;
            padding: 0;
            background: none;
            border: none;
            color: #333;
            display: flex;
            gap: 6mm;
        }
        .doc-infobar strong { font-weight: 700; color: #111; margin-right: 1mm; }

        /* ══════════════════════════════════════
           HAUPT-PLANTABELLE
        ══════════════════════════════════════ */
        .plan-table {
            width: 100%;
            border-collapse: collapse;
            font-size: 9pt;
            border: 0.3mm solid #999;
        }
        .plan-table thead th {
            border: 0.3mm solid #999;
            padding: 2mm 3mm;
            text-align: left;
            font-weight: 700;
            background: none;
            color: #111;
            font-size: 9pt;
        }
        .plan-table tbody td {
            border: 0.3mm solid #bbb;
            padding: 1.5mm 3mm;
            vertical-align: top;
        }
        .plan-table tr.day-start td { border-top: 1.2mm solid #222; }
        .plan-table td.liefertag-cell {
            font-weight: 700;
            width: 22mm;
            white-space: nowrap;
            vertical-align: top;
            color: #111;
            text-decoration: underline;
        }
        .plan-table td.bestelltag-cell { width: 24mm; white-space: nowrap; }
        .plan-table td.zeit-cell {
            width: 26mm;
            white-space: nowrap;
        }

        /* ══════════════════════════════════════
           COVER / SEPARATOR
        ══════════════════════════════════════ */
        .separator-page {
            width: 210mm;
            min-height: 297mm;
            margin: 0 auto 28px auto;
            background: #fff;
            box-shadow: 0 8px 32px rgba(0,0,0,0.28);
            padding: 20mm 16mm;
            display: flex;
            flex-direction: column;
            justify-content: center;
            align-items: center;
            text-align: center;
            border-radius: 3px;
        }
        .separator-page h1 { font-size: 26pt; color: #003366; margin-bottom: 8mm; }
        .separator-page h2 { font-size: 15pt; color: #333; margin-bottom: 4mm; }
        .separator-page p   { font-size: 10pt; color: #666; margin: 1mm 0; }

        /* ══════════════════════════════════════
           SUCHE / HIGHLIGHT
        ══════════════════════════════════════ */
        .customer-entry { display: block; contain: layout style; }
        .paper.is-match  { box-shadow: 0 0 0 3px var(--accent), 0 8px 32px rgba(0,0,0,0.28); }
        .paper.is-current { box-shadow: 0 0 0 3px var(--red), 0 8px 32px rgba(0,0,0,0.28); }

        /* ══════════════════════════════════════
           DRUCK
        ══════════════════════════════════════ */
        @page { size: A4 portrait; margin: 0; }

        @media print {
            html, body {
                background: white !important;
                width: 210mm !important;
                margin: 0 !important; padding: 0 !important;
                display: block !important;
            }
            .sidebar, .mobile-toggle { display: none !important; }
            .main-content { width: 210mm !important; }
            .page-stack { padding: 0 !important; }

            .customer-entry { page-break-after: always; break-after: page; }
            .customer-entry:last-child { page-break-after: auto; break-after: auto; }

            .paper {
                width: 210mm !important; height: 297mm !important;
                overflow: hidden !important;
                margin: 0 !important; padding: 0 !important;
                box-shadow: none !important; border-radius: 0 !important;
                page-break-inside: avoid; contain: layout style;
            }
            .paper-inner {
                width: 210mm !important; padding: 10mm 13mm !important;
                box-sizing: border-box !important; transform: none !important;
            }
            .doc-subtitle:hover, .doc-subtitle:focus { background: none; box-shadow: none; }
            .is-match, .is-current { box-shadow: none !important; }
            .print-hidden { display: none !important; }
        }

        /* ══════════════════════════════════════
           RESPONSIVE – Tablets / kleine Screens
        ══════════════════════════════════════ */
        @media screen and (max-width: 768px) {
            .sidebar {
                position: fixed;
                left: -260px;
                transition: left 0.25s ease;
                z-index: 200;
            }
            .sidebar.mobile-open { left: 0; }
            .mobile-toggle {
                display: block;
                position: fixed;
                top: 10px; left: 10px;
                z-index: 201;
                background: var(--sb-active, #e6a100);
                color: #fff;
                border: none; border-radius: 8px;
                padding: 8px 12px;
                font-size: 18px;
                cursor: pointer;
                box-shadow: 0 2px 8px rgba(0,0,0,0.2);
            }
            .main-content { margin-left: 0 !important; }
        }
        @media screen and (min-width: 769px) {
            .mobile-toggle { display: none; }
        }

        /* ══════════════════════════════════════
           CONTENTEDITABLE DIRTY-STATE
        ══════════════════════════════════════ */
        .doc-subtitle.is-edited {
            background: rgba(200,0,0,0.06);
            border-bottom: 2px dashed rgba(200,0,0,0.3);
        }

        /* ══════════════════════════════════════
           QUELLDATEN-PANEL (rechte Sideview)
        ══════════════════════════════════════ */
        .src-btn {
            background: #eef4ff; border: 1.5px solid #93b4e0; border-radius: 6px;
            padding: 3px 12px; font-size: 10px; font-weight: 700;
            color: #1a60b0; cursor: pointer; font-family: inherit;
            transition: all 0.15s; margin-left: auto;
        }
        .src-btn:hover { background: #dbeafe; border-color: #1a60b0; color: #0f4a8a; }
        @media print { .src-btn { display: none !important; } }

        .src-overlay {
            position: fixed; inset: 0; background: rgba(0,0,0,0.35);
            z-index: 400; display: none;
        }
        .src-overlay.open { display: flex; justify-content: flex-end; }
        .src-panel {
            width: min(680px, 92vw); height: 100vh;
            background: #fff; box-shadow: -8px 0 40px rgba(0,0,0,0.18);
            display: flex; flex-direction: column;
            overflow: hidden;
        }
        .src-panel-header {
            padding: 16px 20px 12px; background: #f8fafc;
            border-bottom: 1px solid #e5e9f0;
            display: flex; align-items: center; justify-content: space-between;
            flex-shrink: 0;
        }
        .src-panel-title { font-size: 14px; font-weight: 700; color: #1a2332; }
        .src-panel-close {
            background: none; border: none; font-size: 18px; color: #6b7a90;
            cursor: pointer; padding: 4px 8px; border-radius: 6px;
        }
        .src-panel-close:hover { background: #e8ecf2; color: #1a2332; }
        .src-panel-body {
            flex: 1; overflow-y: auto; padding: 16px 20px;
            scrollbar-width: thin; scrollbar-color: #dde2ea transparent;
        }
        .src-table {
            width: 100%; border-collapse: collapse; font-size: 11px;
            margin-bottom: 12px;
        }
        .src-table thead th {
            background: #f0f4f8; color: #4a5568; padding: 6px 8px;
            text-align: left; font-size: 10px; font-weight: 700;
            letter-spacing: 0.05em; border-bottom: 1.5px solid #dde2ea;
            white-space: nowrap;
        }
        .src-table tbody td {
            padding: 5px 8px; border-bottom: 1px solid #f0f4f8; color: #2a3848;
        }
        .src-table tbody tr:hover td { background: #f5f7fa; }

        /* ══════════════════════════════════════
           ANFAHRTSHINWEIS-BOX
        ══════════════════════════════════════ */
        .anfahrt-box {
            margin-top: 3mm;
            border: 0.3mm solid #bbb;
            border-left: 1.2mm solid #e6a100;
            border-radius: 1.5mm;
            padding: 2mm 3mm;
            font-size: 8.5pt;
            color: #333;
            background: #fffdf5;
            line-height: 1.5;
        }
        .anfahrt-box-label {
            font-size: 7.5pt;
            font-weight: 700;
            text-transform: uppercase;
            letter-spacing: 0.08em;
            color: #b07800;
            margin-bottom: 1mm;
        }
        .anfahrt-meta {
            display: flex;
            gap: 6mm;
            font-size: 8pt;
            color: #555;
            margin-top: 1.5mm;
            padding-top: 1.5mm;
            border-top: 0.2mm solid #e0d8c0;
        }
        .anfahrt-meta strong { color: #333; margin-right: 1mm; }
    </style>
    """


def render_tour_overview(customer_rows: pd.DataFrame) -> str:
    """Baut die Tourübersicht: zeigt vorhandene Liefertage."""
    if customer_rows.empty:
        return ""

    day_order = ["Montag", "Dienstag", "Mittwoch", "Donnerstag", "Freitag", "Samstag", "Sonntag"]

    all_days = set(
        d for d in customer_rows["Liefertag"].dropna().unique()
        if d and d != "Unbekannt"
    )

    if not all_days:
        return ""

    days_present = [d for d in day_order if d in all_days]
    n_cols = len(days_present)
    label_w = "18mm"
    col_w = f"calc((100% - {label_w}) / {n_cols})"
    day_spans = "".join(
        f'<span style="display:inline-block;width:{col_w}">{html.escape(d)}</span>' for d in days_present
    )

    return f"""
    <div style="font-size:9pt; margin-bottom:2.5mm; line-height:1.6;">
        <div><strong style="display:inline-block;width:{label_w}">Liefertag:</strong>{day_spans}</div>
    </div>
    """


def render_plan_table(rows: pd.DataFrame) -> str:
    """Haupttabelle: Liefertag (rowspan) | Sortiment | Bestelltag | Bestellzeitende.
    Entspricht exakt dem PDF-Vorbild: Liefertag-Zelle nur einmal pro Tag, fett."""
    if rows.empty:
        return "<p>Keine Planzeilen vorhanden.</p>"

    day_order = ["Montag", "Dienstag", "Mittwoch", "Donnerstag", "Freitag", "Samstag", "Sonntag", "Unbekannt"]

    # Sortieren: erst nach Liefertag (Wochentag-Reihenfolge), dann Sortiment
    def day_sort_key(day: str) -> int:
        try:
            return day_order.index(day)
        except ValueError:
            return 99

    def time_to_minutes(t: str) -> int:
        """'20:00' -> 1200, '09:15' -> 555, '' -> 9999"""
        try:
            clean = t.replace(" Uhr", "").strip()
            h, m = clean.split(":")
            return int(h) * 60 + int(m)
        except Exception:
            return 9999

    ordered = rows.copy()
    ordered["_day_order"]  = ordered["Liefertag"].map(day_sort_key)
    ordered["_time_order"] = ordered["Bestellzeitende"].map(normalize_text).map(time_to_minutes)
    ordered = ordered.sort_values(["_day_order", "SortKey_Sortiment", "_time_order"], ascending=[True, True, False])

    # Rowspan pro Liefertag zählen + HTML in einem Pass
    _rec_cols = ["Liefertag", "Sortiment", "Bestelltag_Name", "Bestellzeitende"]
    if "Liefertyp_ID" in ordered.columns:
        _rec_cols.append("Liefertyp_ID")
    records = ordered[_rec_cols].fillna("").to_dict("records")

    # Erst Counts ermitteln (schneller Vorlauf)
    day_counts: dict = {}
    for rec in records:
        d = rec["Liefertag"] or "Unbekannt"
        day_counts[d] = day_counts.get(d, 0) + 1

    body_rows: list = []
    day_seen: set = set()

    for rec in records:
        day        = rec["Liefertag"] or "Unbekannt"
        tg_id      = str(rec.get("Liefertyp_ID", "")).strip()
        sortiment  = rec["Sortiment"]
        # Transportgruppen-Nummer vor Sortiment anzeigen
        sort_display = f"{tg_id} – {sortiment}" if tg_id and sortiment else sortiment or tg_id
        bestelltag = rec["Bestelltag_Name"]
        zeitende   = rec["Bestellzeitende"]

        # "Uhr" anhängen falls nicht schon vorhanden
        if zeitende and "uhr" not in zeitende.lower():
            zeitende = zeitende + " Uhr"

        is_day_start = day not in day_seen
        day_seen.add(day)

        day_cell = ""
        if is_day_start:
            rowspan = day_counts[day]
            day_cell = f'<td class="liefertag-cell" rowspan="{rowspan}">{html.escape(day)}</td>'

        tr_class = ' class="day-start"' if is_day_start else ""

        body_rows.append(
            f"""<tr{tr_class}>
                {day_cell}
                <td class="sortiment-cell">{html.escape(sort_display)}</td>
                <td class="bestelltag-cell">{html.escape(bestelltag)}</td>
                <td class="zeit-cell">{html.escape(zeitende)}</td>
            </tr>"""
        )

    return f"""
    <table class="plan-table">
        <thead>
            <tr>
                <th style="width:20mm;">Liefertag</th>
                <th>Sortiment</th>
                <th style="width:22mm;">Bestelltag</th>
                <th style="width:24mm;">Bestellzeitende</th>
            </tr>
        </thead>
        <tbody>
            {"".join(body_rows)}
        </tbody>
    </table>
    """


def logo_img_tag(logo_b64: str, logo_mime: str = "image/png") -> str:
    """Gibt ein <img>-Tag zurück. Im Bulk-Export wird src per JS gesetzt (einmalig im Head),
    im Einzel-Export direkt als Data-URI. Beide tragen class=doc-logo-img.
    """
    if logo_b64:
        # Einzeldokument: src direkt; Bulk: wird per JS-Injection überschrieben (kein doppeltes Einbetten)
        return (
            f'<img class="doc-logo-img" '
            f'src="data:{logo_mime};base64,{logo_b64}" '
            f'alt="NORDfrische Center" '
            f'style="max-width:44mm; max-height:20mm; width:auto; height:auto; display:block; margin-left:auto;">'
        )
    # CSS-Fallback ohne Logo
    return """
        <div style="display:inline-flex; align-items:flex-start; gap:3px;">
            <div style="border:1.5px solid #003366; padding:2mm 3mm; line-height:1.25; display:inline-block;">
                <span style="display:block; font-size:6.5pt; font-weight:800; color:#003366;
                             letter-spacing:0.06em; text-transform:uppercase;">NORDfrische</span>
                <span style="display:block; font-size:9.5pt; font-weight:900; color:#003366;">Center</span>
                <span style="display:block; font-size:5.5pt; color:#555;">Das Fleischwerk von EDEKA Nord</span>
            </div>
            <div style="border:1.5px solid #f5a623; background:#f5a623; color:#003366;
                        font-size:12pt; font-weight:900; padding:1.5mm 2.5mm; line-height:1.15;">NORD</div>
        </div>"""


def _logo_bulk_placeholder() -> str:
    """Im Bulk-Export: leeres img-Tag, src wird per JS aus window.LOGO_DATA gesetzt."""
    return (
        '<img class="doc-logo-img" src="" alt="NORDfrische Center" '
        'style="max-width:44mm; max-height:20mm; width:auto; height:auto; display:block; margin-left:auto;">'
    )


def render_customer_plan(
    customer: pd.Series,
    customer_rows: pd.DataFrame,
    logo_b64: str = "",
    logo_mime: str = "image/png",
    bulk_mode: bool = False,
    photo_map: Optional[Dict[str, str]] = None,
) -> str:
    """Rendert eine einzelne Kundenseite exakt nach dem PDF-Vorbild.
    Hinweis: Kundendaten sind bereits in cleanup_dataframe() normalisiert,
    daher kein erneuter normalize_text-Aufruf nötig.
    """
    sap_nr      = str(customer.get("SAP_Nr", ""))
    name        = str(customer.get("Name", ""))
    strasse     = str(customer.get("Strasse", ""))
    plz         = str(customer.get("PLZ", ""))
    ort         = str(customer.get("Ort", ""))
    fachberater = str(customer.get("Fachberater", ""))
    anfahrt_text = str(customer.get("Anfahrtshinweis", "")).strip()
    behaelter    = str(customer.get("Behaelter", "")).strip()
    sap_nr_csv   = str(customer.get("SAP_Nr_CSV", "")).strip()
    stand = datetime.now().strftime("%d.%m.%Y")

    subtitle = "Standard"  # Immer Standard – per contenteditable änderbar

    tour_overview_html = render_tour_overview(customer_rows)
    plan_table_html    = render_plan_table(customer_rows)

    # Anfahrtshinweis-Box nur wenn Inhalt vorhanden
    if anfahrt_text or behaelter or sap_nr_csv:
        meta_parts = []
        if sap_nr_csv:
            meta_parts.append(f'<span><strong>SAP-Nr.:</strong> {html.escape(sap_nr_csv)}</span>')
        if behaelter:
            meta_parts.append(f'<span><strong>Behälter:</strong> {html.escape(behaelter)}</span>')
        meta_html = (
            f'<div class="anfahrt-meta">{"".join(meta_parts)}</div>'
            if meta_parts else ""
        )
        # Foto einbetten falls vorhanden
        foto_filename = str(customer.get("Foto", "")).strip()
        foto_html = ""
        if foto_filename and photo_map:
            foto_data_uri = photo_map.get(foto_filename.lower(), "")
            if foto_data_uri:
                foto_html = (
                    f'<img src="{foto_data_uri}" alt="{html.escape(foto_filename)}" '
                    f'style="max-width:38mm; max-height:28mm; width:auto; height:auto; '
                    f'border-radius:2mm; border:0.3mm solid #ddd; margin-top:2mm; display:block;">'
                )
        anfahrt_html = f"""
        <div class="anfahrt-box" style="display:flex; gap:3mm; align-items:flex-start;">
            <div style="flex:1; min-width:0;">
                <div class="anfahrt-box-label">&#128205; Anfahrtshinweis</div>
                <div>{html.escape(anfahrt_text)}</div>
                {meta_html}
            </div>
            {f'<div style="flex-shrink:0;">{foto_html}</div>' if foto_html else ''}
        </div>"""
    else:
        anfahrt_html = ""

    return f"""
    <div class="paper">
    <div class="paper-inner">

        <!-- ===== HEADER: Adresse | Titel | Logo ===== -->
        <div class="doc-header">
            <div class="doc-address">
                <strong>{html.escape(name)}</strong><br>
                {html.escape(strasse)}<br>
                {html.escape(plz)} {html.escape(ort)}
            </div>

            <div class="doc-title-block">
                <div class="doc-title">Sende- &amp; Belieferungsplan</div>
                <div class="doc-subtitle" contenteditable="true" title="Klicken zum Bearbeiten">{html.escape(subtitle)}</div>
            </div>

            <div class="doc-logo">
                {_logo_bulk_placeholder() if bulk_mode else logo_img_tag(logo_b64, logo_mime)}
            </div>
        </div>

        <!-- ===== INFOLEISTE ===== -->
        <div class="doc-infobar">
            <span><strong>Kunden-Nr.:</strong> {html.escape(sap_nr)}</span>
            <span><strong>Fachberater:</strong> {html.escape(fachberater)}</span>
            <span><strong>Stand:</strong> {html.escape(stand)}</span>
            <button type="button" class="src-btn" onclick="openSourcePanel('{html.escape(sap_nr)}')">&#128270; Quelldaten</button>
            <span class="md-tour-inline" style="display:none"></span>
        </div>

        <!-- ===== TOUR-ÜBERSICHT ===== -->
        {tour_overview_html}

        <!-- ===== PLANTABELLE ===== -->
        {plan_table_html}

        <!-- ===== ANFAHRTSHINWEIS ===== -->
        {anfahrt_html}

    </div>
    </div>
    """


def render_separator_page(customer: pd.Series) -> str:
    return f"""
    <div class="separator-page">
        <h1>{html.escape(str(customer.get('Name', '')))}</h1>
        <h2>SAP {html.escape(str(customer.get('SAP_Nr', '')))}</h2>
        <p>CSB {html.escape(str(customer.get('CSB_Nr', '')))}</p>
        <p>{html.escape(str(customer.get('PLZ', '')))} {html.escape(str(customer.get('Ort', '')))}</p>
    </div>
    """


def render_export_search_toolbar(massendruck_section: str = "", logo_b64: str = "", logo_mime: str = "image/png") -> str:
    if logo_b64:
        logo_html = (
            f'<img src="data:{logo_mime};base64,{logo_b64}" '
            f'alt="Logo" style="max-width:48px;max-height:48px;width:auto;height:auto;'
            f'border-radius:8px;flex-shrink:0;object-fit:contain;">'
        )
    else:
        logo_html = '<div class="sidebar-logo-icon">&#128230;</div>'

    return f"""
    <button type="button" class="mobile-toggle" id="mobile-toggle"
        onclick="document.getElementById('sidebar').classList.toggle('mobile-open')"
        aria-label="Menü öffnen">&#9776;</button>
    <aside class="sidebar" id="sidebar">
        <div class="sidebar-logo">
            {logo_html}
            <div>
                <div class="sidebar-logo-text">Sende- &amp; Belieferungsplan</div>
                <div class="sidebar-logo-sub">NORDfrische Center</div>
            </div>
        </div>

        <div class="sidebar-section">
            <div class="sidebar-label">Suche</div>
            <input id="search-input" type="text"
                placeholder="SAP, CSB, Name, Ort \u2026"
                autocomplete="off" spellcheck="false" />
            <div class="search-nav-row">
                <button type="button" class="search-btn" id="btn-prev" title="Vorheriger (Shift+Enter)">&#8679;</button>
                <button type="button" class="search-btn" id="btn-next" title="N\u00e4chster (Enter)">&#8681;</button>
                <button type="button" class="search-btn reset" id="btn-reset" title="Zur\u00fccksetzen (Esc)">&#10005;</button>
                <span class="search-count" id="search-count"></span>
            </div>
            <div class="search-empty" id="search-empty">Keine Treffer.</div>
        </div>

        <div class="sidebar-section">
            <div class="sidebar-label">Kunden</div>
            <span class="search-count" id="cnt-alle" style="font-size:12px;color:#4a5568"></span>
        </div>

        <div class="sidebar-subtitle-group">
            <div class="sidebar-label">Untertitel global \u00e4ndern</div>
            <input id="global-subtitle-input" type="text"
                placeholder="z.B. Standard, NMS \u2026"
                autocomplete="off" spellcheck="false" />
            <div class="search-nav-row" style="margin-top:6px;">
                <button type="button" class="search-btn" id="btn-apply-subtitle">&#10003; Alle setzen</button>
            </div>
        </div>

        """ + massendruck_section + """

        <button type="button" class="sidebar-print-btn" onclick="printCurrent()">&#128438; Drucken</button>
        <button type="button" class="sidebar-debug-btn" onclick="toggleDebug()">&#128269; Debug</button>
    </aside>
    """


def _build_debug_html(data: Optional[Dict[str, pd.DataFrame]]) -> str:
    if not data:
        return ""
    sections = []
    all_csv_parts: List[str] = []

    for title, df in data.items():
        count = len(df)
        icon = "✅" if count == 0 else "⚠️"
        if df.empty:
            rows_html = "<tr><td colspan='99' style='color:#888;padding:8px'>Keine Einträge</td></tr>"
            thead_html = ""
            export_btn = ""
        else:
            cols = list(df.columns)
            thead_html = "<thead><tr>" + "".join(f"<th>{html.escape(c)}</th>" for c in cols) + "</tr></thead>"
            rows_html_parts: List[str] = []
            csv_lines: List[str] = [";".join(cols)]
            for rec in df.to_dict("records"):
                row_cells = [str(rec[c]) for c in cols]
                rows_html_parts.append("<tr>" + "".join(
                    f"<td>{html.escape(cell)}</td>" for cell in row_cells
                ) + "</tr>")
                csv_lines.append(";".join(f'"{cell}"' for cell in row_cells))
            rows_html = "".join(rows_html_parts)

            csv_bytes = "\n".join(csv_lines).encode("utf-8-sig")
            csv_b64 = base64.b64encode(csv_bytes).decode()
            safe_title = title.replace("/", "-").replace(" ", "_")
            export_btn = (
                f'<a class="dbg-export" '
                f'href="data:text/csv;base64,{csv_b64}" '
                f'download="debug_{html.escape(safe_title)}.csv">&#8595; CSV</a>'
            )

            all_csv_parts.append(f"=== {title} ===")
            all_csv_parts.extend(csv_lines)
            all_csv_parts.append("")

        sections.append(f"""
        <div class="dbg-section">
            <div class="dbg-title" onclick="this.parentElement.classList.toggle('open')">
                <span>{icon} {html.escape(title)}</span>
                <div style="display:flex;align-items:center;gap:8px;">
                    <span class="dbg-count">{count}</span>
                    {export_btn}
                </div>
            </div>
            <div class="dbg-body">
                <table class="dbg-table">
                    {thead_html}
                    <tbody>{rows_html}</tbody>
                </table>
            </div>
        </div>""")

    if all_csv_parts:
        all_bytes = "\n".join(all_csv_parts).encode("utf-8-sig")
        all_b64 = base64.b64encode(all_bytes).decode()
        gesamt_btn = (
            f'<a class="dbg-gesamt-export" '
            f'href="data:text/csv;base64,{all_b64}" '
            f'download="sendeplan_debug_gesamt.csv">&#8595; Alle exportieren</a>'
        )
        sections.insert(0, f'<div style="padding:0 0 12px 0;">{gesamt_btn}</div>')

    return "".join(sections)


def _rows_to_list(df: pd.DataFrame, cols: List[str]) -> list:
    avail = [c for c in cols if c in df.columns]
    return df[avail].fillna("").astype(str).to_dict(orient="records")


def build_full_document_html(customers: pd.DataFrame, plan_rows: pd.DataFrame, include_separators: bool = False, logo_b64: str = "", logo_mime: str = "image/png", sidebar_logo_b64: str = "", sidebar_logo_mime: str = "image/png", debug_data: Optional[Dict[str, pd.DataFrame]] = None, massendruck_data: Optional[dict] = None, photo_map: Optional[Dict[str, str]] = None) -> str:
    # Logo einmalig als JS-Variable – wird nach DOMContentLoaded auf alle Bilder gesetzt.
    # Spart mehrere MB bei größeren Kundenstämmen (logo_b64 × N Kunden).
    if logo_b64:
        logo_head_script = f"""
        <script>
        (function() {{
            var src = "data:{logo_mime};base64,{logo_b64}";
            document.addEventListener("DOMContentLoaded", function() {{
                document.querySelectorAll(".doc-logo-img").forEach(function(img) {{ img.src = src; }});
            }});
        }})();
        </script>"""
    else:
        logo_head_script = ""


    # ── Massendruck: JSON-Daten + Sidebar-Sektion + JS aufbauen ──
    if massendruck_data:
        md_json = json.dumps(massendruck_data, ensure_ascii=False)
        md_days_json = json.dumps({str(k): v for k, v in WOCHENTAGE.items()}, ensure_ascii=False)
        massendruck_data_script = f"""
        <script>
        window.MASSENDRUCK = {{
            assignments: {md_json},
            days: {md_days_json}
        }};
        </script>"""

        massendruck_sidebar_section = """
        <div class="sidebar-section md-section" id="md-section">
            <div class="sidebar-label">&#128438; Massendruck &ndash; Sortiert auf Normalwoche</div>
            <div class="md-day-row" id="md-day-row">
                <button class="md-day-btn" data-day="1">Mo</button>
                <button class="md-day-btn" data-day="2">Di</button>
                <button class="md-day-btn" data-day="3">Mi</button>
                <button class="md-day-btn" data-day="4">Do</button>
                <button class="md-day-btn" data-day="5">Fr</button>
                <button class="md-day-btn" data-day="6">Sa</button>
            </div>
            <div class="md-stats" id="md-stats" style="display:none"></div>
            <div class="md-btn-row" id="md-btn-row" style="display:none">
                <button type="button" class="md-overview-btn" onclick="openMdOverlay()">
                    &#128269; Reihenfolge ansehen
                </button>
                <button type="button" class="sidebar-print-btn md-print-btn"
                    onclick="printMassendruck()">&#128438; Drucken</button>
            </div>
        </div>

        <!-- Massendruck Overlay -->
        <div id="md-overlay" class="md-overlay" style="display:none" onclick="if(event.target===this)closeMdOverlay()">
            <div class="md-overlay-box">
                <div class="md-overlay-header">
                    <div class="md-overlay-title" id="md-overlay-title">Druckreihenfolge</div>
                    <button class="md-overlay-close" onclick="closeMdOverlay()">&#10005;</button>
                </div>
                <div class="md-overlay-stats" id="md-overlay-stats"></div>
                <div class="md-overlay-table-wrap">
                    <table class="md-table">
                        <thead><tr>
                            <th style="width:36px">#</th>
                            <th>Kundenname</th>
                            <th style="width:54px">SAP-Nr</th>
                            <th id="md-th-p" style="width:72px">Prim\u00e4r</th>
                            <th id="md-th-s" style="width:72px">Sekund\u00e4r</th>
                            <th style="width:70px">Priorit\u00e4t</th>
                        </tr></thead>
                        <tbody id="md-table-body"></tbody>
                    </table>
                </div>
                <div class="md-overlay-footer">
                    <button type="button" class="sidebar-print-btn" style="width:auto;padding:10px 28px"
                        onclick="printMassendruck()">&#128438; Drucken</button>
                    <button type="button" class="md-overview-btn" onclick="closeMdOverlay()">Schlie\u00dfen</button>
                </div>
            </div>
        </div>"""

        massendruck_css = """
        /* ── Massendruck Sidebar (helles Design) ── */
        .md-day-row {
            display: flex; gap: 4px; flex-wrap: wrap; margin-bottom: 6px;
        }
        .md-day-btn {
            flex: 1; min-width: 28px;
            border: 1.5px solid var(--sb-border, #dde2ea); border-radius: 6px;
            padding: 5px 2px; font-size: 11px; font-weight: 600;
            font-family: inherit; cursor: pointer;
            background: #fff; color: #4a5568;
            transition: all 0.14s;
        }
        .md-day-btn:hover { background: var(--sb-hover, #e8ecf2); color: #1a2332; border-color: #b0bac8; }
        .md-day-btn.active { background: var(--sb-active, #e6a100); color: #fff; border-color: transparent;
            box-shadow: 0 2px 6px rgba(230,161,0,0.3); }
        .md-stats {
            font-size: 10px; line-height: 1.8; margin-bottom: 6px;
            padding: 6px 9px; background: #fff;
            border-radius: 6px; border: 1px solid var(--sb-border, #dde2ea);
            color: #3a4a5c;
        }
        .md-btn-row {
            display: flex; flex-direction: column; gap: 5px;
        }
        .md-overview-btn {
            display: block; width: 100%;
            border: 1.5px solid var(--sb-border, #dde2ea); border-radius: 8px;
            padding: 8px 10px; font-size: 12px; font-weight: 600;
            font-family: inherit; cursor: pointer;
            background: #fff; color: #2a3848;
            text-align: center; transition: all 0.14s;
        }
        .md-overview-btn:hover { background: var(--sb-hover, #e8ecf2); border-color: #b0bac8; }
        .md-print-btn { margin: 0 !important; width: 100% !important; font-size: 12px !important; }

        /* ── Overlay (bleibt dunkel – Modal über weißen Papieren) ── */
        .md-overlay {
            position: fixed; inset: 0;
            background: rgba(0,0,0,0.6);
            z-index: 500;
            display: flex; align-items: center; justify-content: center;
            padding: 20px;
        }
        .md-overlay-box {
            background: #fff;
            border: 1px solid #dde2ea;
            border-radius: 14px;
            width: min(900px, 96vw);
            max-height: 88vh;
            display: flex; flex-direction: column;
            box-shadow: 0 24px 80px rgba(0,0,0,0.2);
            overflow: hidden;
        }
        .md-overlay-header {
            display: flex; align-items: center; justify-content: space-between;
            padding: 18px 22px 14px;
            border-bottom: 1px solid #e5e9f0;
            background: #f8fafc;
            flex-shrink: 0;
        }
        .md-overlay-title {
            font-size: 15px; font-weight: 700; color: #1a2332;
        }
        .md-overlay-close {
            background: none; border: none; color: #6b7a90; font-size: 18px;
            cursor: pointer; padding: 4px 8px; border-radius: 6px;
            transition: all 0.15s;
        }
        .md-overlay-close:hover { background: #e8ecf2; color: #1a2332; }
        .md-overlay-stats {
            padding: 10px 22px 8px;
            font-size: 12px; line-height: 1.8; color: #4a5568;
            flex-shrink: 0;
            border-bottom: 1px solid #e5e9f0;
            background: #fafbfc;
        }
        .md-overlay-table-wrap {
            flex: 1; overflow-y: auto; overflow-x: auto;
            scrollbar-width: thin; scrollbar-color: #dde2ea transparent;
        }
        .md-overlay-footer {
            padding: 14px 22px;
            border-top: 1px solid #e5e9f0;
            background: #f8fafc;
            display: flex; gap: 10px; align-items: center;
            flex-shrink: 0;
        }
        .md-table {
            width: 100%; border-collapse: collapse; font-size: 12px;
        }
        .md-table thead th {
            position: sticky; top: 0;
            background: #f0f4f8; color: #4a5568; padding: 8px 12px;
            text-align: left; font-size: 10px; letter-spacing: 0.06em;
            font-weight: 700; text-transform: uppercase;
            border-bottom: 1.5px solid #dde2ea;
            white-space: nowrap;
        }
        .md-table tbody td {
            padding: 7px 12px;
            border-bottom: 1px solid #f0f4f8;
            color: #2a3848;
        }
        .md-table tbody tr:hover td { background: #f5f7fa; }
        .md-table .md-tour { font-family: 'Courier New', monospace; font-size: 11px; font-weight: 600; }
        .md-prio-p { color: #1a7f3c; font-weight: 700; }
        .md-prio-s { color: #1a60b0; font-weight: 700; }
        .md-prio-u { color: #aab2be; }
        @media print { .md-section, .md-overlay { display: none !important; } }
        .md-tour-inline {
            font-family: 'Courier New', monospace;
            font-size: 9pt;
            font-weight: 700;
            color: #1a5a9e;
            letter-spacing: 0.06em;
            margin-left: auto;
        }
        @media screen { .md-tour-inline { display: none !important; } }
        """

        massendruck_js = """
        // ── Massendruck ──
        (function() {
            if (!window.MASSENDRUCK) return;
            var MD = window.MASSENDRUCK;
            var activeMdDay = null;
            var lastOrdered = [];

            function escHtml(s) {
                return (s||'').replace(/&/g,'&amp;').replace(/</g,'&lt;').replace(/>/g,'&gt;');
            }

            function computeOrder(primaryDay) {
                var entries = window._allEntries || Array.from(document.querySelectorAll('.customer-entry'));
                var ordered = entries.map(function(entry) {
                    var sap    = (entry.getAttribute('data-sap') || '').trim();
                    var name   = entry.getAttribute('data-name') || '';
                    var asgn   = MD.assignments[sap] || {};
                    var pt     = asgn[String(primaryDay)] || '';

                    // Pro Kunde: alle restlichen Tage zyklisch durchsuchen bis Tour gefunden
                    var st = ''; var stDay = null;
                    if (!pt) {
                        for (var i = 1; i <= 5; i++) {
                            var candidate = ((primaryDay - 1 + i) % 6) + 1;
                            if (asgn[String(candidate)]) {
                                st    = asgn[String(candidate)];
                                stDay = candidate;
                                break;
                            }
                        }
                    }

                    var prio       = pt ? 0 : (st ? 1 : 2);
                    var tourDigits = (pt || st || '').replace(/\\D/g,'').padStart(8,'0');
                    return {
                        entry: entry, pt: pt, st: st, stDay: stDay,
                        prio: prio, name: name,
                        sap: entry.getAttribute('data-sap') || '',
                        key: prio + tourDigits + name
                    };
                });
                ordered.sort(function(a,b){ return a.key < b.key ? -1 : a.key > b.key ? 1 : 0; });
                return ordered;
            }

            // Erstellt Statistik-HTML für die Sidebar
            function buildStatsHtml(ordered, pdName) {
                var pCount = 0, uCount = 0;
                var sByDay = {};   // { dayNum: { name, count } }
                ordered.forEach(function(o) {
                    if (o.prio === 0) { pCount++; }
                    else if (o.prio === 1) {
                        var d = String(o.stDay);
                        if (!sByDay[d]) sByDay[d] = { name: MD.days[d] || ('Tag ' + d), count: 0 };
                        sByDay[d].count++;
                    } else { uCount++; }
                });
                var html = '<span style="color:#1a7f3c">&#9679; Prim\u00e4r (' + escHtml(pdName) + '): <strong>' + pCount + '</strong></span><br>';
                var sDays = Object.keys(sByDay).sort(function(a,b){ return parseInt(a)-parseInt(b); });
                sDays.forEach(function(d) {
                    var sName = sByDay[d].name;
                    html += '<span style="color:#1a60b0">&nbsp;&nbsp;&#8627; Sekund\u00e4r ' + escHtml(sName) + ': <strong>' + sByDay[d].count + '</strong></span><br>';
                });
                html += '<span style="color:#9a9a9a">&#9679; Keine Tour: <strong>' + uCount + '</strong></span>';
                return html;
            }

            function buildTable(ordered, pdName) {
                var thP = document.getElementById('md-th-p');
                var thS = document.getElementById('md-th-s');
                if (thP) thP.textContent = pdName.slice(0,2) + '-Tour (Prim\u00e4r)';
                if (thS) thS.textContent = 'Sekund\u00e4r-Tour (Tag)';

                var tbody = document.getElementById('md-table-body');
                tbody.innerHTML = '';
                ordered.forEach(function(o, i) {
                    var prioLabel = o.prio === 0
                        ? '<span class="md-prio-p">Prim\u00e4r</span>'
                        : o.prio === 1
                            ? '<span class="md-prio-s">Sek. ' + escHtml(MD.days[String(o.stDay)] || '') + '</span>'
                            : '<span class="md-prio-u">\u00dcbrig</span>';
                    var stCell = o.st ? (escHtml(o.st) + ' <span style="color:#aab2be;font-size:10px">(' + escHtml((MD.days[String(o.stDay)]||'').slice(0,2)) + ')</span>') : '';
                    var tr = document.createElement('tr');
                    tr.innerHTML =
                        '<td style="color:#6b7a90;text-align:right;padding-right:8px">' + (i+1) + '</td>' +
                        '<td style="font-weight:600;color:#1a2332">' + escHtml(o.name) + '</td>' +
                        '<td style="font-family:monospace;font-size:11px;color:#6b7a90">' + escHtml(o.sap) + '</td>' +
                        '<td class="md-tour" style="color:#b07800">' + escHtml(o.pt) + '</td>' +
                        '<td class="md-tour">' + stCell + '</td>' +
                        '<td>' + prioLabel + '</td>';
                    tbody.appendChild(tr);
                });
                return ordered.length;
            }

            function applyMassendruck(primaryDay) {
                activeMdDay = primaryDay;
                var pdName = MD.days[String(primaryDay)] || ('Tag ' + primaryDay);

                lastOrdered = computeOrder(primaryDay);

                // DOM-Reihenfolge anpassen
                var stack = document.querySelector('.page-stack');
                lastOrdered.forEach(function(o) { stack.appendChild(o.entry); });
                window._allEntries = lastOrdered.map(function(o) { return o.entry; });

                // Tour-Nummer in Infoleiste (Druckansicht) eintragen
                lastOrdered.forEach(function(o) {
                    var span = o.entry.querySelector('.md-tour-inline');
                    if (!span) return;
                    var tour = o.pt || o.st || '';
                    span.style.display = '';
                    span.textContent = tour ? 'Tour: ' + tour : '';
                });

                var stats = document.getElementById('md-stats');
                stats.style.display = '';
                stats.innerHTML = buildStatsHtml(lastOrdered, pdName) +
                    '<br><span style="color:#9a6800;font-size:9px">' + lastOrdered.length + ' Kunden</span>';

                var row = document.getElementById('md-btn-row');
                if (row) row.style.display = '';

                // Tag-Buttons
                document.querySelectorAll('.md-day-btn').forEach(function(b) {
                    b.classList.toggle('active', parseInt(b.getAttribute('data-day')) === primaryDay);
                });
            }

            window.openMdOverlay = function() {
                if (activeMdDay === null) return;
                var primaryDay = activeMdDay;
                var pdName   = MD.days[String(primaryDay)] || ('Tag ' + primaryDay);

                var title = document.getElementById('md-overlay-title');
                if (title) title.textContent = 'Druckreihenfolge \u2013 Prim\u00e4rtag: ' + pdName;

                var nr = buildTable(lastOrdered, pdName);

                var ostats = document.getElementById('md-overlay-stats');
                if (ostats) ostats.innerHTML = buildStatsHtml(lastOrdered, pdName) +
                    '&nbsp;&nbsp;<span style="color:#9a6800">' + nr + ' Kunden</span>';

                var overlay = document.getElementById('md-overlay');
                if (overlay) overlay.style.display = 'flex';
            };

            window.closeMdOverlay = function() {
                var overlay = document.getElementById('md-overlay');
                if (overlay) overlay.style.display = 'none';
            };

            window.printMassendruck = function() {
                closeMdOverlay();
                window.print();
            };

            document.addEventListener('DOMContentLoaded', function() {
                document.querySelectorAll('.md-day-btn').forEach(function(btn) {
                    btn.addEventListener('click', function() {
                        applyMassendruck(parseInt(btn.getAttribute('data-day')));
                    });
                });
                document.addEventListener('keydown', function(e) {
                    if (e.key === 'Escape') closeMdOverlay();
                });
            });
        })();
        """
    else:
        massendruck_data_script = ""
        massendruck_sidebar_section = ""
        massendruck_css = ""
        massendruck_js = ""

    debug_html = _build_debug_html(debug_data)
    docs_buffer = io.StringIO()
    search_index: List[str] = []  # Kompaktes JSON-Array statt data-search-Attribut pro Kunde

    # Vorab gruppieren statt pro Kunde den gesamten DataFrame zu filtern
    _plan_grouped = {sap: grp for sap, grp in plan_rows.groupby("SAP_Nr")}

    entry_count = 0
    for _, customer in customers.iterrows():
        sap = customer.get("SAP_Nr", "")
        rows = _plan_grouped.get(sap, pd.DataFrame(columns=plan_rows.columns)).copy()
        csb_nr = customer.get("CSB_Nr", "")
        # Volltext-Blob: alle durchsuchbaren Felder zusammenfassen
        sortimente_text = " ".join(sorted({
            normalize_text(v) for v in rows.get("Sortiment", pd.Series(dtype=str)).tolist() if normalize_text(v)
        }))
        search_blob = " ".join(
            part for part in [
                sap, csb_nr,
                customer.get("Name", ""),
                customer.get("Ort", ""),
                customer.get("PLZ", ""),
                customer.get("Strasse", ""),
                customer.get("Fachberater", ""),
                sortimente_text,
            ]
            if part
        ).lower()
        search_index.append(search_blob)

        entry_parts: List[str] = []
        if include_separators:
            entry_parts.append(render_separator_page(customer))
        entry_parts.append(render_customer_plan(customer, rows, logo_b64="", logo_mime=logo_mime, bulk_mode=True, photo_map=photo_map))

        cust_name_escaped = html.escape(str(customer.get("Name", "")).lower())
        docs_buffer.write(
            f'<section class="customer-entry" '
            f'data-idx="{entry_count}" '
            f'data-sap="{html.escape(sap.lower())}" '
            f'data-csb="{html.escape(csb_nr.lower())}" '
            f'data-name="{cust_name_escaped}">'
            f'{"".join(entry_parts)}'
            f'</section>'
        )
        entry_count += 1

    # Quelldaten pro Kunde (SAP vs KSP) als JSON für die Sideview
    source_data: Dict[str, dict] = {}
    _src_cols = ["Liefertag", "Sortiment", "Bestelltag_Name", "Bestellzeitende", "KSP_Schluessel"]
    for sap_nr, grp in _plan_grouped.items():
        ist_zusatz = grp["_ist_zusatz"].astype(bool) if "_ist_zusatz" in grp.columns else pd.Series(False, index=grp.index)
        sap_rows = grp[~ist_zusatz]
        ksp_rows = grp[ist_zusatz]
        source_data[str(sap_nr).lower()] = {
            "sap": _rows_to_list(sap_rows, _src_cols),
            "ksp": _rows_to_list(ksp_rows, _src_cols),
        }

    source_data_script = (
        '<script>window._sourceData='
        + json.dumps(source_data, ensure_ascii=False)
        + ';</script>'
    )

    source_panel_js = """
    <script>
    function openSourcePanel(sap) {
        function esc(s) { return (s||'').replace(/&/g,'&amp;').replace(/</g,'&lt;').replace(/>/g,'&gt;'); }
        var dayOrder = {'Montag':1,'Dienstag':2,'Mittwoch':3,'Donnerstag':4,'Freitag':5,'Samstag':6,'Sonntag':7};
        var SD = window._sourceData || {};
        var data = SD[sap.toLowerCase()];
        var title = document.getElementById('src-panel-title');
        var body  = document.getElementById('src-panel-body');
        title.textContent = 'Quelldaten \\u2013 SAP ' + sap;
        if (!data) { body.innerHTML = '<p style="color:#888">Keine Daten.</p>'; }
        else {
            var days = {};
            data.sap.forEach(function(r) {
                var d = r.Liefertag || 'Unbekannt';
                if (!days[d]) days[d] = [];
                days[d].push({src:'SAP', sort:r.Sortiment||'', btag:r.Bestelltag_Name||'', bzeit:r.Bestellzeitende||'', ksp:r.KSP_Schluessel||''});
            });
            data.ksp.forEach(function(r) {
                var d = r.Liefertag || 'Unbekannt';
                if (!days[d]) days[d] = [];
                days[d].push({src:'CSB', sort:r.Sortiment||'', btag:r.Bestelltag_Name||'', bzeit:r.Bestellzeitende||'', ksp:r.KSP_Schluessel||''});
            });
            var sorted = Object.keys(days).sort(function(a,b) {
                return (dayOrder[a]||99) - (dayOrder[b]||99);
            });

            var h = '';
            sorted.forEach(function(day) {
                var rows = days[day];
                rows.sort(function(a,b) {
                    var pinWords = ['fleisch','heidemark'];
                    var al = a.sort.toLowerCase(), bl = b.sort.toLowerCase();
                    var ap = pinWords.some(function(w){return al.indexOf(w)!==-1;});
                    var bp = pinWords.some(function(w){return bl.indexOf(w)!==-1;});
                    if (ap !== bp) return ap ? -1 : 1;
                    var sa = a.src === 'SAP' ? 0 : 1;
                    var sb = b.src === 'SAP' ? 0 : 1;
                    if (sa !== sb) return sa - sb;
                    if (a.bzeit > b.bzeit) return -1;
                    if (a.bzeit < b.bzeit) return 1;
                    return 0;
                });
                h += '<div style="margin:18px 0 8px;font-size:13px;font-weight:800;color:#1a2332;border-bottom:2px solid #e5e9f0;padding-bottom:5px">';
                h += esc(day) + '</div>';
                h += '<table class="src-table"><thead><tr>';
                h += '<th style="width:42px">Quelle</th><th>Sortiment</th><th>Bestelltag</th><th>Bestellzeit</th><th>KSP-Key</th>';
                h += '</tr></thead><tbody>';
                rows.forEach(function(r) {
                    if (r.src === 'CSB') {
                        h += '<tr style="background:#eefbf0;border-left:3px solid #1a9e52">';
                        h += '<td style="font-size:9px;font-weight:800;color:#1a7f3c;letter-spacing:0.08em">CSB</td>';
                        h += '<td style="font-weight:600;color:#1a7f3c">' + esc(r.sort) + '</td>';
                    } else {
                        h += '<tr style="background:#f8faff;border-left:3px solid #4a90d9">';
                        h += '<td style="font-size:9px;font-weight:800;color:#1a60b0;letter-spacing:0.08em">SAP</td>';
                        h += '<td>' + esc(r.sort) + '</td>';
                    }
                    h += '<td>' + esc(r.btag) + '</td><td>' + esc(r.bzeit) + '</td>';
                    h += '<td style="font-family:monospace;color:#6b7a90;font-size:10px">' + esc(r.ksp) + '</td></tr>';
                });
                h += '</tbody></table>';
            });
            body.innerHTML = h;
        }
        document.getElementById('src-overlay').classList.add('open');
    }
    function closeSourcePanel() { document.getElementById('src-overlay').classList.remove('open'); }
    document.addEventListener('keydown', function(e) { if (e.key === 'Escape') closeSourcePanel(); });
    </script>
    """

    # Suchdaten als kompaktes JSON-Array – spart ~100KB HTML bei 500 Kunden
    search_data_script = (
        '<script>window._searchData='
        + json.dumps(search_index, ensure_ascii=False)
        + ';</script>'
    )

    search_script = """
    <script>
    (function () {
        "use strict";
        var allEntries = [];
        var matches    = [];
        var cursor     = -1;

        function norm(s) {
            return (s || "").toLowerCase()
                .normalize("NFD").replace(/[\u0300-\u036f]/g, "")
                .replace(/[^a-z0-9 ]/g, " ")
                .replace(/ +/g, " ").trim();
        }

        function papers(entry) {
            return entry.querySelectorAll(".paper");
        }

        function setClass(entry, cls, on) {
            papers(entry).forEach(function (p) {
                p.classList.toggle(cls, on);
            });
        }

        function clearHighlights() {
            matches.forEach(function (e) {
                setClass(e, "is-match",   false);
                setClass(e, "is-current", false);
            });
        }

        function scrollToCurrent() {
            if (cursor < 0 || cursor >= matches.length) return;
            var entry = matches[cursor];
            setClass(entry, "is-current", true);
            entry.scrollIntoView({ behavior: "smooth", block: "start" });
        }

        function updateCounts() {
            var q = norm(document.getElementById("search-input").value);
            var SD = window._searchData || [];
            var total = 0;
            allEntries.forEach(function (e) {
                var idx = parseInt(e.getAttribute("data-idx"), 10);
                var blob = norm(SD[idx] || "");
                if (!q || blob.indexOf(q) !== -1) total++;
            });
            var el = document.getElementById("cnt-alle");
            if (el) el.textContent = total + " Kunden";
        }

        function updateSearchCount() {
            var lbl = document.getElementById("search-count");
            var emp = document.getElementById("search-empty");
            var q   = document.getElementById("search-input").value.trim();
            if (!q) {
                var vis = allEntries.filter(function (e) { return e.style.display !== "none"; }).length;
                lbl.textContent = vis + " Kunden";
                emp.style.display = "none";
            } else {
                if (matches.length === 0) {
                    lbl.textContent = "0 Treffer";
                    emp.style.display = "block";
                } else {
                    lbl.textContent = (cursor + 1) + " / " + matches.length;
                    emp.style.display = "none";
                }
            }
        }

        function applyFilter() {
            var q = norm(document.getElementById("search-input").value);
            var SD = window._searchData || [];
            clearHighlights();
            matches = [];
            cursor  = -1;

            allEntries.forEach(function (entry) {
                var idx  = parseInt(entry.getAttribute("data-idx"), 10);
                var blob = norm(SD[idx] || "");
                var srchOk = !q || blob.indexOf(q) !== -1;
                entry.style.display = srchOk ? "" : "none";
                if (srchOk && q) {
                    setClass(entry, "is-match", true);
                    matches.push(entry);
                }
            });

            if (matches.length > 0) {
                cursor = 0;
                setClass(matches[0], "is-current", true);
                scrollToCurrent();
            }
            updateSearchCount();
            updateCounts();
        }

        function step(dir) {
            if (matches.length === 0) return;
            setClass(matches[cursor], "is-current", false);
            cursor = (cursor + dir + matches.length) % matches.length;
            setClass(matches[cursor], "is-current", true);
            scrollToCurrent();
            updateSearchCount();
        }

        function resetSearch() {
            clearHighlights();
            document.getElementById("search-input").value = "";
            allEntries.forEach(function (e) { e.style.display = ""; });
            matches = [];
            cursor  = -1;
            updateSearchCount();
            updateCounts();
        }

        document.addEventListener("DOMContentLoaded", function () {
            allEntries = Array.from(document.querySelectorAll(".customer-entry"));
            window._allEntries = allEntries;  // für Massendruck-IIFE sichtbar

            // Debounce: bei schnellem Tippen nicht bei jedem Tastendruck filtern
            var _searchTimer = null;
            document.getElementById("search-input").addEventListener("input", function () {
                clearTimeout(_searchTimer);
                _searchTimer = setTimeout(function () { applyFilter(); }, 150);
            });
            document.getElementById("btn-next").addEventListener("click",  function () { step(1); });
            document.getElementById("btn-prev").addEventListener("click",  function () { step(-1); });
            document.getElementById("btn-reset").addEventListener("click", resetSearch);

            document.getElementById("search-input").addEventListener("keydown", function (e) {
                if (e.key === "Enter")  { e.preventDefault(); step(e.shiftKey ? -1 : 1); }
                if (e.key === "Escape") { resetSearch(); }
            });

            // Globaler Untertitel
            function applyGlobalSubtitle() {
                var val = document.getElementById("global-subtitle-input").value;
                if (!val.trim()) return;
                document.querySelectorAll(".doc-subtitle").forEach(function (el) {
                    el.textContent = val;
                    el.classList.add("is-edited");
                });
            }
            document.getElementById("btn-apply-subtitle").addEventListener("click", applyGlobalSubtitle);
            document.getElementById("global-subtitle-input").addEventListener("keydown", function (e) {
                if (e.key === "Enter") { e.preventDefault(); applyGlobalSubtitle(); }
            });

            // Contenteditable dirty-state: visuelles Feedback bei manueller Bearbeitung
            document.querySelectorAll(".doc-subtitle").forEach(function (el) {
                var orig = el.textContent;
                el.addEventListener("input", function () {
                    el.classList.toggle("is-edited", el.textContent !== orig);
                });
            });

            updateSearchCount();
            updateCounts();

            // ── Lazy fitToPage: nur sichtbare Papers skalieren (IntersectionObserver) ──
            var supportsZoom = 'zoom' in document.documentElement.style &&
                !/firefox/i.test(navigator.userAgent);

            function fitSinglePaper(inner) {
                // Reset
                inner.style.zoom = "";
                inner.style.transform = "";
                inner.style.transformOrigin = "";
                inner.style.width = "210mm";
                // Measure
                var paperH = inner.parentElement.clientHeight;
                var paperW = inner.parentElement.clientWidth;
                var contentH = inner.scrollHeight;
                var contentW = inner.scrollWidth;
                var scale = Math.min(paperH / contentH, paperW / contentW, 1);
                if (scale < 1) {
                    if (supportsZoom) {
                        inner.style.zoom = scale;
                    } else {
                        inner.style.transform = "scale(" + scale + ")";
                        inner.style.transformOrigin = "top left";
                    }
                    inner.style.width = "210mm";
                }
            }

            // Lazy: nur Papers skalieren wenn sie in den Viewport kommen
            var fitObserver = new IntersectionObserver(function (entries) {
                entries.forEach(function (entry) {
                    if (entry.isIntersecting) {
                        var inner = entry.target.querySelector(".paper-inner");
                        if (inner) fitSinglePaper(inner);
                        fitObserver.unobserve(entry.target);
                    }
                });
            }, { rootMargin: "200px 0px" });

            document.querySelectorAll(".paper").forEach(function (p) {
                fitObserver.observe(p);
            });

            var _resizeTimer = null;
            window.addEventListener("resize", function () {
                clearTimeout(_resizeTimer);
                _resizeTimer = setTimeout(function () {
                    // Bei Resize alle sichtbaren Papers neu skalieren
                    document.querySelectorAll(".paper").forEach(function (p) {
                        fitObserver.observe(p);
                    });
                }, 200);
            });

            // ── Aktuell sichtbaren Kunden tracken (IntersectionObserver) ──
            var currentVisible = null;
            var visObserver = new IntersectionObserver(function (entries) {
                entries.forEach(function (entry) {
                    if (entry.isIntersecting) {
                        currentVisible = entry.target;
                    }
                });
            }, { threshold: 0.3 });
            allEntries.forEach(function (e) { visObserver.observe(e); });
        });

        window.printCurrent = function printCurrent() {
            // Wenn Suche aktiv und genau 1 Treffer: diesen drucken
            // Sonst: aktuell sichtbaren Kunden im Viewport drucken
            var entries = Array.from(document.querySelectorAll(".customer-entry"));
            var visible = entries.filter(function (e) { return e.style.display !== "none"; });

            var target = null;
            if (visible.length === 1) {
                target = visible[0];
            } else if (typeof currentVisible !== "undefined" && currentVisible) {
                target = currentVisible;
            }

            if (!target) {
                window.print();
                return;
            }

            // Alle anderen verstecken, drucken, wiederherstellen
            entries.forEach(function (e) {
                if (e !== target) e.classList.add("print-hidden");
            });
            window.print();
            entries.forEach(function (e) { e.classList.remove("print-hidden"); });
        };
    })();
    """ + massendruck_js + """
    </script>
    """

    return f"""
    <!DOCTYPE html>
    <html lang="de">
    <head>
        <meta charset="utf-8" />
        <meta name="viewport" content="width=device-width, initial-scale=1.0" />
        <link rel="icon" href="data:image/svg+xml,<svg xmlns='http://www.w3.org/2000/svg' viewBox='0 0 100 100'><text y='.9em' font-size='90'>📦</text></svg>">
        <title>Sendeplan-Export</title>
        {export_css()}
        <style>
        .sidebar-debug-btn {{
            display:block; width:calc(100% - 28px); margin:0 14px 10px;
            border:1.5px solid var(--sb-border,#dde2ea); border-radius:10px;
            padding:9px; font-size:12px; font-weight:600; font-family:inherit;
            cursor:pointer; background:#fff; color:#4a5568;
            text-align:center; transition:all 0.15s;
        }}
        .sidebar-debug-btn:hover {{ background:var(--sb-hover,#e8ecf2); color:#1a2332; border-color:#b0bac8; }}
        @media print {{ .sidebar-debug-btn,.debug-panel {{ display:none !important; }} }}
        .debug-panel {{
            display:none; position:fixed; top:0; right:0; bottom:0;
            width:720px; max-width:92vw;
            background:#111b25; border-left:1px solid rgba(255,255,255,0.1);
            z-index:200; overflow-y:auto; padding:20px;
            font-family:'Segoe UI',system-ui,sans-serif;
            box-shadow:-8px 0 32px rgba(0,0,0,0.5);
        }}
        .debug-panel.open {{ display:block; }}
        .debug-panel-header {{
            display:flex; align-items:center; justify-content:space-between;
            margin-bottom:16px; padding-bottom:12px;
            border-bottom:1px solid rgba(255,255,255,0.1);
        }}
        .debug-panel-title {{ font-size:15px; font-weight:700; color:#fff; }}
        .debug-close {{
            background:none; border:none; color:#aaa; font-size:20px;
            cursor:pointer; padding:4px 8px; border-radius:6px;
        }}
        .debug-close:hover {{ background:rgba(255,255,255,0.1); color:#fff; }}
        .dbg-section {{
            margin-bottom:8px; border:1px solid rgba(255,255,255,0.08);
            border-radius:8px; overflow:hidden;
        }}
        .dbg-title {{
            display:flex; justify-content:space-between; align-items:center;
            padding:10px 14px; background:rgba(255,255,255,0.05);
            cursor:pointer; font-size:12px; font-weight:600; color:#ccc;
            user-select:none;
        }}
        .dbg-title:hover {{ background:rgba(255,255,255,0.09); }}
        .dbg-count {{
            font-family:'Courier New',monospace;
            background:rgba(255,255,255,0.1);
            padding:2px 8px; border-radius:20px; font-size:11px;
        }}
        .dbg-body {{ display:none; overflow-x:auto; }}
        .dbg-section.open .dbg-body {{ display:block; }}
        .dbg-table {{
            width:100%; border-collapse:collapse; font-size:11px; color:#ccc;
        }}
        .dbg-table thead th {{
            background:#0d2035; color:#fff; padding:6px 8px; text-align:left;
            font-size:10px; letter-spacing:0.05em;
            border-bottom:1px solid rgba(255,255,255,0.1);
        }}
        .dbg-table tbody td {{
            padding:5px 8px; border-bottom:1px solid rgba(255,255,255,0.05);
        }}
        .dbg-table tbody tr:hover td {{ background:rgba(255,255,255,0.04); }}
        .dbg-export {{
            font-size:10px; font-weight:600; color:#f0a500;
            text-decoration:none; padding:2px 8px;
            border:1px solid rgba(240,165,0,0.3);
            border-radius:5px; white-space:nowrap;
        }}
        .dbg-export:hover {{ background:rgba(240,165,0,0.15); }}
        .dbg-gesamt-export {{
            display:inline-block; font-size:12px; font-weight:700;
            color:#0d2035; background:var(--accent,#f0a500);
            text-decoration:none; padding:7px 14px;
            border-radius:8px; margin-bottom:4px;
        }}
        .dbg-gesamt-export:hover {{ opacity:0.9; }}
        {massendruck_css}
        </style>
        <script>
        function toggleDebug() {{ document.getElementById('debug-panel').classList.toggle('open'); }}
        document.addEventListener('click', function(e) {{
            if (e.target.classList.contains('dbg-export') || e.target.classList.contains('dbg-gesamt-export')) {{
                e.stopPropagation();
            }}
        }});
        </script>
        {massendruck_data_script}
        {logo_head_script}
        {search_data_script}
        {source_data_script}
    </head>
    <body>
        {render_export_search_toolbar(massendruck_sidebar_section, logo_b64=sidebar_logo_b64, logo_mime=sidebar_logo_mime)}
        <div class="main-content">
        <div class="page-stack">
        {docs_buffer.getvalue()}
        </div>
        </div>
        <div class="debug-panel" id="debug-panel">
            <div class="debug-panel-header">
                <div class="debug-panel-title">&#128269; SAP Debug</div>
                <button class="debug-close" onclick="toggleDebug()">&#10005;</button>
            </div>
            {debug_html if debug_html else '<p style="color:#666;font-size:12px">Keine Debug-Daten vorhanden.</p>'}
        </div>
        <!-- Quelldaten-Panel -->
        <div class="src-overlay" id="src-overlay" onclick="if(event.target===this)closeSourcePanel()">
            <div class="src-panel">
                <div class="src-panel-header">
                    <div class="src-panel-title" id="src-panel-title">Quelldaten</div>
                    <button class="src-panel-close" onclick="closeSourcePanel()">&#10005;</button>
                </div>
                <div class="src-panel-body" id="src-panel-body"></div>
            </div>
        </div>
        {source_panel_js}
        {search_script}
    </body>
    </html>
    """


def init_session_state() -> None:
    pass  # Session-State wird bei Bedarf in main() gesetzt


def all_required_uploads_present(upload_map: Dict[str, Optional[object]]) -> bool:
    return all(upload_map.values())


def show_customer_preview(customer: pd.Series, customer_rows: pd.DataFrame) -> None:
    st.markdown(f"### {customer['Name']}")
    st.caption(f"SAP {customer['SAP_Nr']} · CSB {customer['CSB_Nr']} · {customer['PLZ']} {customer['Ort']}")

    info_cols = st.columns(4)
    info_cols[0].metric("SAP-Nummer", normalize_text(customer.get("SAP_Nr", "")) or "-")
    info_cols[1].metric("CSB-Nummer", normalize_text(customer.get("CSB_Nr", "")) or "-")
    info_cols[2].metric("Fachberater", normalize_text(customer.get("Fachberater", "")) or "-")
    info_cols[3].metric("Planzeilen", len(customer_rows))

    address = " · ".join(
        part for part in [
            normalize_text(customer.get("Strasse", "")),
            " ".join(filter(None, [normalize_text(customer.get("PLZ", "")), normalize_text(customer.get("Ort", ""))])),
        ]
        if part
    )
    st.write(f"**Adresse:** {address or '-'}")

    # Anfahrtsdaten anzeigen falls vorhanden
    anfahrt = normalize_text(customer.get("Anfahrtshinweis", ""))
    behaelter = normalize_text(customer.get("Behaelter", ""))
    sap_nr_csv = normalize_text(customer.get("SAP_Nr_CSV", ""))
    if anfahrt or behaelter or sap_nr_csv:
        with st.expander("📍 Anfahrtshinweis", expanded=True):
            if anfahrt:
                st.write(anfahrt)
            meta = []
            if sap_nr_csv:
                meta.append(f"**SAP-Nr. (CSV):** {sap_nr_csv}")
            if behaelter:
                meta.append(f"**Behälter:** {behaelter}")
            if meta:
                st.caption("  ·  ".join(meta))

    st.markdown("#### Planliste")
    if customer_rows.empty:
        st.warning("Für diesen Kunden sind aktuell keine Planzeilen vorhanden.")
        return

    table = customer_rows.sort_values(["SortKey_Bestelltag", "SortKey_Sortiment", "Bestellzeitende"], ascending=[True, True, False]).copy()
    table = table[["Liefertag", "Sortiment", "Bestelltag_Name", "Bestellzeitende"]].rename(
        columns={
            "Sortiment": "Eintrag",
            "Bestelltag_Name": "Bestelltag",
        }
    )
    st.dataframe(table, use_container_width=True, hide_index=True)


def main() -> None:
    init_session_state()
    st.markdown(streamlit_css(), unsafe_allow_html=True)

    # ── App-Header ──
    h_left, h_right = st.columns([1, 3], gap="medium")
    with h_left:
        app_logo_file = st.file_uploader(
            "App-Logo",
            type=["png", "jpg", "jpeg", "svg", "gif", "webp"],
            key="app_logo",
            help="Logo für die App-Kopfzeile (wird nicht gedruckt)",
        )
        if app_logo_file is not None:
            st.image(app_logo_file, use_container_width=True)
        else:
            st.markdown(
                """
                <div style="border:2px dashed #30363d; border-radius:10px; padding:18px 12px;
                            text-align:center; color:#888; font-size:0.8rem; line-height:1.5;">
                    📷<br>App-Logo<br><span style="font-size:0.7rem;color:#555">PNG · JPG · SVG</span>
                </div>
                """,
                unsafe_allow_html=True,
            )
    with h_right:
        st.markdown(
            """
            <div style="padding: 6px 0 2px 0;">
                <div style="font-size:1.75rem; font-weight:800; color:#e0e0e0; line-height:1.15;
                            letter-spacing:-0.01em;">
                    Sende- &amp; Belieferungsplan
                </div>
                <div style="font-size:1.05rem; font-weight:500; color:#f0a500; margin-top:3px;
                            letter-spacing:0.04em;">
                    NORDfrische Center
                </div>
                <div style="font-size:0.78rem; color:#666; margin-top:6px;">
                    Sendeplan-Generator · EDEKA Nord Fleisch
                </div>
            </div>
            """,
            unsafe_allow_html=True,
        )

    st.divider()

    # ── Uploads ──
    col_left, col_right = st.columns(2, gap="medium")
    with col_left:
        kunden_file = st.file_uploader("Kundenliste", type=["xlsx", "xls", "xlsm", "csv"],
                                        help="Spalten: A, I, J, K, L, M, N")
        sap_file = st.file_uploader("SAP-Datei", type=["xlsx", "xls", "xlsm", "csv"],
                                     help="Spalten: A, G, H, I, O, P, Y")
        anfahrt_file = st.file_uploader(
            "Anfahrtshinweise (optional)",
            type=["csv", "xlsx", "xls", "xlsm"],
            key="anfahrt_upload",
            help="A=SAP-Nr · B=CSB-Nr (Matchkey) · G=Behälter · H=Anfahrtshinweis · K=Foto-Dateiname",
        )
    with col_right:
        transport_file = st.file_uploader("Transportgruppen", type=["xlsx", "xls", "xlsm", "csv"],
                                          help="Spalten: A, C")
        kostenstellen_file = st.file_uploader("Kostenstellen-Datei", type=["xlsx", "xls", "xlsm", "csv"],
                                              help="A=Liefertag, B=Tourname, dann Sortiment-Gruppen (Lagerware, AVO, …)")
        foto_files = st.file_uploader(
            "Kunden-Fotos (optional, mehrere)",
            type=["jpg", "jpeg", "png", "webp", "gif"],
            accept_multiple_files=True,
            key="foto_upload",
            help="Dateiname muss Spalte K der Anfahrts-CSV entsprechen (z.B. 40.jpg)",
        )
    logo_file = st.file_uploader(
        "Druck-Logo (Sendeplan)",
        type=["png", "jpg", "jpeg", "svg", "gif", "webp"],
        key="print_logo",
        help="Logo oben rechts auf jedem gedruckten Sendeplan (unabhängig vom App-Logo)",
    )

    upload_map = {
        "kunden": kunden_file, "sap": sap_file, "transport": transport_file,
        "kostenstellen": kostenstellen_file,
    }

    # ── CSV-Trennzeichen ──
    csv_separator = st.selectbox(
        "CSV-Trennzeichen",
        options=[";", ",", "\t"],
        format_func=lambda x: {";" : "Semikolon ;", ",": "Komma ,", "\t": "Tab ⇥"}[x],
        index=0,
        help="Nur relevant für CSV-Uploads. Excel-Dateien ignorieren diese Einstellung.",
    )

    # ── Status-Zeile ──
    file_names = [f'<span class="status-ok">✓ {html.escape(v.name)}</span>' if v else '<span class="status-miss">✗ fehlt</span>'
                  for k, v in upload_map.items()]
    labels = ["Kunden", "SAP", "Transport", "Kostenstellen"]
    status_parts = [f"{l}: {f}" for l, f in zip(labels, file_names)]
    st.markdown(f"<p style='font-size:0.85rem;margin:0.5rem 0;'>{'&ensp;·&ensp;'.join(status_parts)}</p>", unsafe_allow_html=True)

    if not all_required_uploads_present(upload_map):
        st.info("Alle 4 Dateien hochladen, dann erscheint der Button.")
        return

    # ── Foto-Map aufbauen: {dateiname_lower -> data_uri} ──
    photo_map: Dict[str, str] = {}
    for _f in (foto_files or []):
        _ext = _f.name.rsplit(".", 1)[-1].lower()
        _mime = {"jpg": "image/jpeg", "jpeg": "image/jpeg", "png": "image/png",
                 "webp": "image/webp", "gif": "image/gif"}.get(_ext, "image/jpeg")
        _b64 = base64.b64encode(_f.getvalue()).decode("utf-8")
        photo_map[_f.name.lower()] = f"data:{_mime};base64,{_b64}"

    # ── Daten verarbeiten (csv_separator kommt aus Selectbox oben) ──
    try:
        _hasher = hashlib.md5()
        for _f in (kunden_file, sap_file, transport_file, kostenstellen_file):
            _hasher.update(_f.getvalue())
        if anfahrt_file:
            _hasher.update(anfahrt_file.getvalue())
        for _f in (foto_files or []):
            _hasher.update(_f.name.encode())
            _hasher.update(_f.getvalue())
        _hasher.update(csv_separator.encode())
        _cache_key = _hasher.hexdigest()

        if st.session_state.get("_df_cache_key") != _cache_key:
            _result = prepare_dataframes(
                kunden_file.getvalue(), kunden_file.name,
                sap_file.getvalue(), sap_file.name,
                transport_file.getvalue(), transport_file.name,
                kostenstellen_file.getvalue(), kostenstellen_file.name,
                csv_separator,
                anfahrt_bytes=anfahrt_file.getvalue() if anfahrt_file else None,
                anfahrt_name=anfahrt_file.name if anfahrt_file else None,
            )
            st.session_state["_df_cache_key"] = _cache_key
            st.session_state["_df_cache_result"] = _result
            st.session_state["_export_ready"] = False  # alte HTML verwerfen

        (customers_df, plan_rows_df, counts,
         df_sap_debug) = st.session_state["_df_cache_result"]
    except Exception as exc:
        st.error(f"Fehler beim Verarbeiten: {exc}")
        return

    # Debug-Reports cachen
    _data_key = st.session_state.get("_df_cache_key", "")
    if st.session_state.get("_debug_cache_key") != _data_key:
        st.session_state["_debug_reports"] = build_debug_report(plan_rows_df, df_sap_debug)
        st.session_state["_debug_cache_key"] = _data_key
    debug_reports = st.session_state["_debug_reports"]

    # ── Logos vorbereiten ──
    # Druck-Logo: erscheint oben rechts auf jedem A4-Sendeplan
    logo_b64 = ""
    logo_mime = "image/png"
    if logo_file is not None:
        logo_b64 = base64.b64encode(logo_file.getvalue()).decode("utf-8")
        ext = logo_file.name.rsplit(".", 1)[-1].lower()
        logo_mime = {"jpg": "image/jpeg", "jpeg": "image/jpeg", "png": "image/png",
                     "svg": "image/svg+xml", "gif": "image/gif", "webp": "image/webp"}.get(ext, "image/png")

    # Sidebar-Logo: erscheint im HTML-Seitenleisten-Header
    sidebar_logo_b64 = ""
    sidebar_logo_mime = "image/png"
    if app_logo_file is not None:
        sidebar_logo_b64 = base64.b64encode(app_logo_file.getvalue()).decode("utf-8")
        ext2 = app_logo_file.name.rsplit(".", 1)[-1].lower()
        sidebar_logo_mime = {"jpg": "image/jpeg", "jpeg": "image/jpeg", "png": "image/png",
                             "svg": "image/svg+xml", "gif": "image/gif", "webp": "image/webp"}.get(ext2, "image/png")

    st.divider()

    # ── Tabs: Plan | Vorschau | Debug ──
    total_issues = sum(len(df) for df in debug_reports.values())
    debug_label = f"🔍 Debug ({total_issues} ⚠️)" if total_issues > 0 else "🔍 Debug ✅"
    tab_plan, tab_preview, tab_debug = st.tabs(["⚡ Export", "👁 Kundenvorschau", debug_label])

    # ── Tab: Export ──
    with tab_plan:
        st.markdown(
            f"**{len(customers_df)} Kunden** · {len(plan_rows_df)} Planzeilen"
        )

        include_sep = st.checkbox(
            "Trennseiten einfügen (Separator-Pages vor jedem Kunden)",
            value=False,
            help="Fügt vor jede Kundenseite ein A4-Deckblatt mit Name und SAP-Nr. ein.",
        )

        if st.button("⚡ Plan generieren", use_container_width=True, type="primary"):
            progress = st.progress(0, text="Vorbereitung …")
            n = len(customers_df)
            # Fortschritt: HTML-Build mit Zwischenmeldungen
            with st.spinner(f"Generiere HTML für {n} Kunden …"):
                bulk_html = build_full_document_html(
                    customers_df, plan_rows_df,
                    include_separators=include_sep,
                    logo_b64=logo_b64, logo_mime=logo_mime,
                    sidebar_logo_b64=sidebar_logo_b64, sidebar_logo_mime=sidebar_logo_mime,
                    debug_data=debug_reports,
                    photo_map=photo_map if photo_map else None,
                )
            progress.progress(100, text="Fertig!")
            st.session_state["_export_html"] = bulk_html
            st.session_state["_export_ready"] = True
            st.toast(f"✅ HTML für {n} Kunden generiert!", icon="📦")

        if st.session_state.get("_export_ready"):
            html_bytes = st.session_state["_export_html"].encode("utf-8")
            size_kb = len(html_bytes) / 1024
            size_label = f"{size_kb/1024:.1f} MB" if size_kb > 1024 else f"{size_kb:.0f} KB"
            st.download_button(
                label=f"⬇  sendeplan.html herunterladen  ({size_label})",
                data=html_bytes,
                file_name="sendeplan.html",
                mime="text/html",
                use_container_width=True,
            )
            st.caption("HTML im Browser öffnen → Suche, Filter, Druck alles drin.")

    # ── Tab: Kundenvorschau ──
    with tab_preview:
        st.markdown("Wähle einen Kunden zur schnellen Voransicht – ohne vollen HTML-Export.")

        search_input = st.text_input(
            "Suche (Name, SAP, CSB, Ort)",
            key="preview_search",
            placeholder="z.B. Edeka Muster oder 1234567",
        )

        filtered = filter_customers(customers_df, search_input)

        if filtered.empty:
            st.warning("Keine Kunden gefunden.")
        else:
            option_labels = {
                row["SAP_Nr"]: f"{row['SAP_Nr']}  |  {row['Name']}  |  {row['Ort']}"
                for _, row in filtered.iterrows()
            }
            selected_sap = st.selectbox(
                f"{len(filtered)} Kunden gefunden – auswählen:",
                options=list(option_labels.keys()),
                format_func=lambda k: option_labels[k],
                key="preview_select",
            )
            if selected_sap:
                customer_row = customers_df[customers_df["SAP_Nr"] == selected_sap].iloc[0]
                customer_plan_rows = plan_rows_df[plan_rows_df["SAP_Nr"] == selected_sap]
                st.divider()
                show_customer_preview(customer_row, customer_plan_rows)

    # ── Tab: Debug ──
    with tab_debug:
        render_debug_tab(debug_reports)


if __name__ == "__main__":
    main()
